#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Excel to YAML Converter for Garmin Planner with Scheduling Support

This module converts a structured Excel file into a YAML file compatible with garmin-planner.
It includes support for a Date column with workout scheduling.
"""

import pandas as pd
import yaml
import re
import os
import sys
import copy
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import argparse
import logging
import random
import string

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Customize YAML dumper to avoid references/aliases
class NoAliasDumper(yaml.SafeDumper):
    """Custom YAML dumper that ignores aliases and supports OrderedDict"""
    def ignore_aliases(self, data):
        return True
    
    def represent_mapping(self, tag, mapping, flow_style=None):
        """Override to handle OrderedDict"""
        # Convert OrderedDict to regular dict for serialization
        if hasattr(mapping, 'items'):
            mapping = dict(mapping.items())
        return super().represent_mapping(tag, mapping, flow_style)

# Valid step types supported by garmin-planner
VALID_STEP_TYPES = {"warmup", "cooldown", "interval", "recovery", "rest", "repeat", "other"}


def extract_heart_rates_from_excel(excel_file):
    """
    Funzione dedicata all'estrazione delle frequenze cardiache da un file Excel.
    Utilizza direttamente openpyxl per leggere le celle.
    
    Args:
        excel_file: Percorso del file Excel
        
    Returns:
        Dizionario contenente le frequenze cardiache estratte
    """
    import openpyxl
    import os
    
    # Verifica che il file esista
    if not os.path.exists(excel_file):
        print(f"File non trovato: {excel_file}")
        return {}
    
    # Dizionario per i risultati
    heart_rates = {}
    
    try:
        # Carica il workbook
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        
        # Verifica se esiste il foglio HeartRates
        if 'HeartRates' not in wb.sheetnames:
            print("Foglio HeartRates non trovato nel file Excel")
            return {}
        
        # Ottieni il foglio
        hr_sheet = wb['HeartRates']
        
        # Stampa le dimensioni per debug
        print(f"Dimensioni foglio HeartRates: {hr_sheet.dimensions}")
        print(f"Numero righe: {hr_sheet.max_row}, Numero colonne: {hr_sheet.max_column}")
        
        # Stampa le prime 10 celle per debug
        for row in range(1, min(10, hr_sheet.max_row + 1)):
            name = hr_sheet.cell(row=row, column=1).value
            value = hr_sheet.cell(row=row, column=2).value
            print(f"Riga {row}: {name} = {value} (tipo: {type(value)})")
        
        # Processa il foglio riga per riga, iniziando dalla seconda riga (indice 2)
        for row in range(2, hr_sheet.max_row + 1):
            # Ottieni il nome e il valore
            name = hr_sheet.cell(row=row, column=1).value
            value = hr_sheet.cell(row=row, column=2).value
            
            # Verifica che entrambi non siano None
            if name is not None and value is not None:
                # Converti il nome in stringa e rimuovi spazi in eccesso
                name_str = str(name).strip()
                
                # Gestisci i vari tipi di valore
                if isinstance(value, (int, float)):
                    heart_rates[name_str] = int(value)
                    print(f"Aggiunta frequenza cardiaca numerica {name_str} = {heart_rates[name_str]}")
                elif isinstance(value, str):
                    heart_rates[name_str] = value.strip()
                    print(f"Aggiunta frequenza cardiaca stringa {name_str} = {heart_rates[name_str]}")
                else:
                    # Per qualsiasi altro tipo, converti in stringa
                    heart_rates[name_str] = str(value).strip()
                    print(f"Aggiunta frequenza cardiaca altro tipo {name_str} = {heart_rates[name_str]}")
        
        # Verifica se abbiamo trovato delle frequenze cardiache
        if heart_rates:
            print(f"Estratte {len(heart_rates)} frequenze cardiache dal foglio HeartRates:")
            for name, value in heart_rates.items():
                print(f"  {name}: {value}")
        else:
            print("Nessuna frequenza cardiaca trovata nel foglio HeartRates")
            
        return heart_rates
    
    except Exception as e:
        import traceback
        print(f"Errore nell'estrazione delle frequenze cardiache: {str(e)}")
        traceback.print_exc()
        return {}

def extract_paces_and_speeds_from_excel(excel_file):
    """
    Estrae ritmi per la corsa, zone di potenza FTP per il ciclismo, passi vasca per il nuoto
    e frequenze cardiache dal file Excel.
    
    Args:
        excel_file: Percorso del file Excel
        
    Returns:
        Tuple (paces, swim_paces, power_values, heart_rates) con i dizionari contenenti i valori estratti
    """
    try:
        import openpyxl
        import re
        import datetime
        
        # Carica il workbook
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        
        # Dizionari da popolare - inizializzati come vuoti invece che usare valori di default
        paces = {}
        swim_paces = {}
        power_values = {}
        heart_rates = {}
        
        # Funzione di supporto per convertire secondi in formato MM:SS
        def seconds_to_mmss(seconds):
            if isinstance(seconds, (int, float)):
                minutes = int(seconds) // 60
                remainder = int(seconds) % 60
                return f"{minutes}:{remainder:02d}"
            return str(seconds)
            
        # Funzione di normalizzazione dei formati di passo
        def normalize_pace_value(value):
            if value is None:
                return None
                
            # Se è un formato 0:MM, convertilo in MM:00
            if isinstance(value, str) and re.match(r'^0:\d{2}$', value):
                minutes = int(value.split(':')[1])
                return f"{minutes}:00"
                
            # Già in formato MM:SS
            if isinstance(value, str) and re.match(r'^\d{1,2}:\d{2}$', value) and not value.startswith('0:'):
                return value
                
            # Se è un formato hh:mm:ss
            if isinstance(value, str) and re.match(r'^\d{1,2}:\d{2}:\d{2}$', value):
                h, m, s = map(int, value.split(':'))
                total_seconds = h * 3600 + m * 60 + s
                minutes = total_seconds // 60
                seconds = total_seconds % 60
                return f"{minutes}:{seconds:02d}"
                
            # Per altri formati
            return seconds_to_mmss(value)
        
        # Estrazione dal foglio Paces
        if 'Paces' in wb.sheetnames:
            sheet = wb['Paces']
            
            # Inizializza lo stato del parser
            current_section = None
            
            # Processa riga per riga
            for row_idx in range(1, sheet.max_row + 1):
                # Leggi i valori dalle colonne
                col0 = sheet.cell(row=row_idx, column=1).value
                col1 = sheet.cell(row=row_idx, column=2).value
                col0_str = str(col0) if col0 is not None else ""
                
                # Debug: stampa la riga
                print(f"Riga {row_idx}: {col0_str} - {col1} ({type(col1)})")
                
                # Controlla se è un'intestazione di sezione
                if col0_str and "RITMI PER LA CORSA" in col0_str.upper():
                    current_section = 'running'
                    print(f"Trovata sezione RUNNING a riga {row_idx}")
                    continue
                elif col0_str and "POTENZA PER IL CICLISMO" in col0_str.upper():
                    current_section = 'power'
                    print(f"Trovata sezione POWER a riga {row_idx}")
                    continue
                elif col0_str and "PASSI VASCA PER IL NUOTO" in col0_str.upper():
                    current_section = 'swimming'
                    print(f"Trovata sezione SWIMMING a riga {row_idx}")
                    continue
                
                # Salta righe vuote, intestazioni o righe di commento
                if not col0 or col0 == "Name" or col0_str.startswith('*') or col0_str.startswith('#'):
                    continue
                
                # Se abbiamo un nome e un valore in una sezione valida
                if current_section and col0 and col1 is not None:
                    # Estrai nome e valore
                    name = col0_str.strip()
                    
                    # Aggiungi al dizionario appropriato con formattazione corretta
                    if current_section == 'running':
                        # Estrai il valore come stringa da qualsiasi tipo
                        value_str = str(col1).strip()
                        
                        # Gestisci il formato specifico 0:MM -> MM:00
                        if re.match(r'^0:\d{2}$', value_str):
                            minutes = int(value_str.split(':')[1])
                            paces[name] = f"{minutes}:00"
                            print(f"Convertito formato 0:MM: {value_str} → {minutes}:00")
                        # Converti in formato MM:SS indipendentemente dal tipo originale
                        elif isinstance(col1, (int, float)):
                            # È già in secondi
                            paces[name] = seconds_to_mmss(col1)
                        elif isinstance(col1, datetime.time):
                            # È un oggetto datetime.time
                            total_seconds = col1.hour * 3600 + col1.minute * 60 + col1.second
                            paces[name] = seconds_to_mmss(total_seconds)
                        elif isinstance(col1, str):
                            # Prova a interpretare come MM:SS o HH:MM:SS
                            if re.match(r'^\d{1,2}:\d{2}$', value_str) and not value_str.startswith('0:'):
                                paces[name] = value_str  # Già in formato MM:SS
                            elif re.match(r'^\d{1,2}:\d{2}:\d{2}$', value_str):
                                parts = value_str.split(':')
                                total_seconds = int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
                                paces[name] = seconds_to_mmss(total_seconds)
                            else:
                                # Prova a convertire in float e poi in MM:SS
                                try:
                                    value_seconds = float(value_str.replace(',', '.'))
                                    paces[name] = seconds_to_mmss(value_seconds)
                                except ValueError:
                                    paces[name] = value_str  # Fallback a stringa
                        else:
                            # Tipo sconosciuto, converti a stringa
                            paces[name] = value_str
                        
                        # Normalizza il valore finale prima di salvarlo
                        paces[name] = normalize_pace_value(paces[name])
                        print(f"Convertito paces {name}: {col1} ({type(col1)}) → {paces[name]}")
                    
                    elif current_section == 'power':
                        # Per potenza, assicurati che sia una stringa
                        power_values[name] = str(col1) if col1 is not None else ""
                    
                    elif current_section == 'swimming':
                        # Estrai il valore come stringa da qualsiasi tipo
                        value_str = str(col1).strip()
                        
                        # Gestisci il formato specifico 0:MM -> MM:00
                        if re.match(r'^0:\d{2}$', value_str):
                            minutes = int(value_str.split(':')[1])
                            swim_paces[name] = f"{minutes}:00"
                            print(f"Convertito formato 0:MM: {value_str} → {minutes}:00")
                        # Stesso processo di conversione per i passi vasca
                        elif isinstance(col1, (int, float)):
                            swim_paces[name] = seconds_to_mmss(col1)
                        elif isinstance(col1, datetime.time):
                            total_seconds = col1.hour * 3600 + col1.minute * 60 + col1.second
                            swim_paces[name] = seconds_to_mmss(total_seconds)
                        elif isinstance(col1, str):
                            if re.match(r'^\d{1,2}:\d{2}$', value_str) and not value_str.startswith('0:'):
                                swim_paces[name] = value_str
                            elif re.match(r'^\d{1,2}:\d{2}:\d{2}$', value_str):
                                parts = value_str.split(':')
                                total_seconds = int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
                                swim_paces[name] = seconds_to_mmss(total_seconds)
                            else:
                                try:
                                    value_seconds = float(value_str.replace(',', '.'))
                                    swim_paces[name] = seconds_to_mmss(value_seconds)
                                except ValueError:
                                    swim_paces[name] = value_str
                        else:
                            swim_paces[name] = value_str
                        
                        # Normalizza il valore finale prima di salvarlo
                        swim_paces[name] = normalize_pace_value(swim_paces[name])
                        print(f"Convertito swim_paces {name}: {col1} ({type(col1)}) → {swim_paces[name]}")
        
        # Estrazione dal foglio HeartRates
        if 'HeartRates' in wb.sheetnames:
            hr_sheet = wb['HeartRates']
            
            # Processa riga per riga, saltando l'intestazione
            for row_idx in range(2, hr_sheet.max_row + 1):
                name = hr_sheet.cell(row=row_idx, column=1).value
                value = hr_sheet.cell(row=row_idx, column=2).value
                
                if name and value is not None:
                    name = str(name).strip()
                    
                    # Gestisci diversi tipi di valori
                    if isinstance(value, (int, float)):
                        heart_rates[name] = int(value)
                    else:
                        heart_rates[name] = str(value).strip()
                    
                    print(f"Aggiunta frequenza cardiaca {name}: {heart_rates[name]}")
        
        # Stampa i valori estratti per debug
        print(f"Ritmi estratti: {paces}")
        print(f"Valori potenza estratti: {power_values}")
        print(f"Passi vasca estratti: {swim_paces}")
        print(f"Frequenze cardiache estratte: {heart_rates}")
        
        return paces, swim_paces, power_values, heart_rates
    
    except Exception as e:
        import traceback
        print(f"Errore nell'estrazione dei ritmi, potenza, passi vasca e frequenze cardiache: {str(e)}")
        traceback.print_exc()
        return {}, {}, {}, {}


def format_pace_for_excel(pace_value):
    """
    Formatta un valore di passo per la visualizzazione nel foglio Excel.
    Assicura che il formato sia sempre MM:SS (mai 0:MM).
    
    Args:
        pace_value: Il valore del ritmo da formattare
        
    Returns:
        Il ritmo formattato correttamente per Excel
    """
    import re
    
    # Se è None o vuoto, ritorna il valore originale
    if pace_value is None or (isinstance(pace_value, str) and not pace_value.strip()):
        return pace_value
    
    # Se è in formato 0:MM (es. '0:06')
    if isinstance(pace_value, str) and re.match(r'^0:\d{2}$', pace_value):
        minutes = int(pace_value.split(':')[1])
        return f"{minutes}:00"
    
    # Se è in formato ssss:00 (es. '380:00' - secondi totali)
    if isinstance(pace_value, str) and re.match(r'^\d+:00$', pace_value):
        try:
            total_seconds = int(pace_value.split(':')[0])
            minutes = total_seconds // 60
            seconds = total_seconds % 60
            return f"{minutes}:{seconds:02d}"
        except (ValueError, IndexError):
            pass
    
    # Se è già in formato standard mm:ss (es. '4:30')
    if isinstance(pace_value, str) and re.match(r'^\d{1,2}:\d{2}$', pace_value):
        return pace_value
    
    # Se è in formato hh:mm:ss (es. '00:04:30')
    if isinstance(pace_value, str) and re.match(r'^\d{1,2}:\d{2}:\d{2}$', pace_value):
        h, m, s = map(int, pace_value.split(':'))
        total_minutes = h * 60 + m
        return f"{total_minutes}:{s:02d}"
    
    # Se è un intero o float (secondi)
    if isinstance(pace_value, (int, float)):
        minutes = int(pace_value) // 60
        seconds = int(pace_value) % 60
        return f"{minutes}:{seconds:02d}"
    
    # Se è un oggetto time
    if hasattr(pace_value, 'hour') and hasattr(pace_value, 'minute') and hasattr(pace_value, 'second'):
        total_seconds = pace_value.hour * 3600 + pace_value.minute * 60 + pace_value.second
        minutes = total_seconds // 60
        seconds = total_seconds % 60
        return f"{minutes}:{seconds:02d}"
    
    # Altre conversioni possono essere aggiunte qui se necessario
    
    return str(pace_value)

def normalize_pace_format(value):
    """
    Normalizza il formato dei ritmi, convertendo vari formati in mm:ss.
    
    Args:
        value: Il valore del ritmo da normalizzare
        
    Returns:
        Il ritmo normalizzato nel formato mm:ss
    """
    import re
    
    # Se è None o vuoto, ritorna il valore originale
    if value is None or (isinstance(value, str) and not value.strip()):
        return value
    
    # Se è già in formato standard mm:ss (es. '4:30')
    if isinstance(value, str) and re.match(r'^\d{1,2}:\d{2}$', value):
        # Gestisci il caso speciale di 0:MM che deve diventare MM:00
        if value.startswith('0:'):
            minutes = int(value.split(':')[1])
            return f"{minutes}:00"
        return value
    
    # Se è in formato hh:mm:ss (es. '00:04:30')
    if isinstance(value, str) and re.match(r'^\d{1,2}:\d{2}:\d{2}$', value):
        h, m, s = map(int, value.split(':'))
        total_minutes = h * 60 + m
        return f"{total_minutes}:{s:02d}"
    
    # Se è in formato 0:MM (es. '0:06')
    if isinstance(value, str) and re.match(r'^0:\d{2}$', value):
        minutes = int(value.split(':')[1])
        return f"{minutes}:00"

    # Se è in formato ssss:00 (es. '380:00' o secondi totali)
    if isinstance(value, str) and re.match(r'^\d+:\d{2}$', value):
        parts = value.split(':')
        if len(parts) == 2:
            try:
                total_seconds = int(parts[0])
                seconds_part = int(parts[1])
                
                # Se i secondi sono 00, interpretiamo come secondi totali
                if seconds_part == 0:
                    minutes = total_seconds // 60
                    seconds = total_seconds % 60
                    return f"{minutes}:{seconds:02d}"
                # Altrimenti manteniamo il formato
                else:
                    return value
            except ValueError:
                # Se non è un numero valido, ritorna il valore originale
                pass
    
    # Se è un numero intero di secondi
    if isinstance(value, (int, float)) or (isinstance(value, str) and value.replace('.', '', 1).isdigit()):
        try:
            total_seconds = int(float(value))
            minutes = total_seconds // 60
            seconds = total_seconds % 60
            return f"{minutes}:{seconds:02d}"
        except (ValueError, TypeError):
            pass
    
    # Se non è riconosciuto, ritorna il valore originale
    return value

def yaml_to_excel(yaml_data, excel_file, create_new=False):
    """
    Converti i dati YAML in un file Excel.
    Mantiene i valori di ritmo, potenza e passi vasca solo nel foglio Paces senza duplicarli in Config.
    
    Args:
        yaml_data: Dizionario con i dati YAML
        excel_file: Percorso del file Excel di output
        create_new: Se True, crea un nuovo file. Se False, aggiorna un file esistente.
    
    Returns:
        True se la conversione è riuscita, False altrimenti
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        import os
        import copy
        import re
        
        # Estrai la configurazione
        config = yaml_data.get('config', {})
        
        sport_type = config.get('sport_type', 'running')
        
        logging.info(f"Tipo di sport rilevato: {sport_type}")
        
        # NUOVO APPROCCIO: Se il file esiste, eliminalo prima di ricrearlo
        if os.path.exists(excel_file):
            try:
                os.remove(excel_file)
                logging.info(f"File Excel esistente eliminato: {excel_file}")
            except PermissionError:
                # Se non possiamo eliminare il file (potrebbe essere aperto), 
                # crea un nuovo file con un nome diverso
                dir_name = os.path.dirname(excel_file)
                base_name = os.path.basename(excel_file)
                name, ext = os.path.splitext(base_name)
                new_file = os.path.join(dir_name, f"{name}_new{ext}")
                logging.warning(f"Impossibile eliminare il file esistente. Creazione di un nuovo file: {new_file}")
                excel_file = new_file
        
        # Crea un nuovo file Excel con il tipo di sport corretto
        create_sample_excel(excel_file, sport_type)
        
        # Carica il nuovo file appena creato
        try:
            wb = openpyxl.load_workbook(excel_file)
        except Exception as e:
            logging.error(f"Errore nel caricamento del file Excel: {str(e)}")
            return False
        
        # Aggiorna la configurazione nel foglio Config (escludendo valori di Paces)
        if 'Config' in wb.sheetnames:
            update_config_sheet(wb['Config'], config)
        
        # Estrai i valori di ritmo, potenza e passi vasca
        # Prima cerca al livello principale, poi in config come fallback (per compatibilità con file vecchi)
        paces = yaml_data.get('paces', {})
        if not paces and 'paces' in config:
            paces = config.get('paces', {})
            
        power_values = yaml_data.get('power_values', {})
        if not power_values and 'power_values' in config:
            power_values = config.get('power_values', {})
            
        swim_paces = yaml_data.get('swim_paces', {})
        if not swim_paces and 'swim_paces' in config:
            swim_paces = config.get('swim_paces', {})
        
        # Aggiorna i ritmi in base al tipo di sport - usando la funzione update_unified_paces_sheet
        if 'Paces' in wb.sheetnames:
            # Rimuovi il foglio Paces esistente e creane uno nuovo
            wb.remove(wb['Paces'])
            paces_sheet = wb.create_sheet('Paces')
            
            # Usa la funzione update_unified_paces_sheet per aggiornare il foglio Paces
            # Questa funzione si occuperà di tutto, compresa la formattazione e la correzione dei valori come 0:06 -> 6:00
            update_unified_paces_sheet(paces_sheet, paces, power_values, swim_paces, sport_type)
        
        # Aggiorna le frequenze cardiache
        if 'HeartRates' in wb.sheetnames:
            heart_rates = {}
            
            # Prima cerca in config
            if 'heart_rates' in config and config['heart_rates']:
                heart_rates = config['heart_rates']
            # Poi cerca a livello radice
            elif 'heart_rates' in yaml_data:
                heart_rates = yaml_data['heart_rates']
            
            # Solo se abbiamo trovato delle frequenze cardiache, aggiorniamo il foglio
            if heart_rates:
                update_heart_rates_sheet(wb['HeartRates'], heart_rates)
            else:
                # Se non ci sono frequenze cardiache nel YAML, pulisci il foglio (mantieni solo l'intestazione)
                hr_sheet = wb['HeartRates']
                for row in range(hr_sheet.max_row, 1, -1):
                    hr_sheet.delete_rows(row)
        
        # Aggiorna gli allenamenti
        if 'Workouts' in wb.sheetnames:
            # CORREZIONE: Crea una deep copy del yaml_data per sicurezza
            yaml_data_copy = copy.deepcopy(yaml_data)
            
            # CORREZIONE: Assicurati che tutte le ripetute siano formattate correttamente
            for key, value in yaml_data_copy.items():
                if isinstance(value, list) and key not in ['config', 'athlete_name', 'paces', 'power_values', 'swim_paces', 'heart_rates']:
                    # Questo è un allenamento, dobbiamo assicurarci che ogni passo repeat sia correttamente formattato
                    for i, step in enumerate(value):
                        if isinstance(step, dict) and 'repeat' in step:
                            # Assicurati che il passo repeat abbia sempre la proprietà steps
                            if 'steps' not in step or not isinstance(step['steps'], list):
                                step['steps'] = []
            
            update_workouts_sheet(wb['Workouts'], yaml_data_copy)
        
        # Garantisci che gli esempi siano sempre presenti
        if 'Examples' in wb.sheetnames:
            wb.remove(wb['Examples'])
        create_unified_examples_sheet(wb)
        
        # Garantisci l'ordine corretto dei fogli
        sheet_order = ['Config', 'Paces', 'HeartRates', 'Workouts', 'Examples']
        wb._sheets = [wb[sheet_name] for sheet_name in sheet_order if sheet_name in wb.sheetnames]
        
        # Salva il file Excel
        try:
            wb.save(excel_file)
            logging.info(f"File Excel salvato con successo: {excel_file}")
            return True
        except Exception as e:
            logging.error(f"Errore nel salvataggio del file Excel: {str(e)}")
            return False
            
    except Exception as e:
        logging.error(f"Errore nella conversione YAML to Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def update_unified_paces_sheet(paces_sheet, paces, power_values, swim_paces, sport_type="running"):
    """
    Aggiorna un foglio Paces unificato con ritmi, potenza e passi vasca.
    
    Args:
        paces_sheet: Foglio Excel per i Paces
        paces: Dizionario con i ritmi per la corsa
        power_values: Dizionario con i valori di potenza per il ciclismo
        swim_paces: Dizionario con i passi vasca per il nuoto
        sport_type: Tipo di sport predefinito
    """
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.cell.cell import TYPE_STRING
    import re
    
    # Pulisci il foglio esistente (rimuovi tutto tranne l'intestazione)
    for row in range(paces_sheet.max_row, 1, -1):
        paces_sheet.delete_rows(row)
    
    # Definisci stili
    header_font = Font(bold=True)
    subheader_font = Font(bold=True, size=12)
    wrapped_alignment = Alignment(wrap_text=True, vertical='top')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    running_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    cycling_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    swimming_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    
    # Imposta le intestazioni di colonna se mancano
    if paces_sheet['A1'].value is None:
        paces_sheet['A1'] = 'Name'
        paces_sheet['B1'] = 'Value'
        paces_sheet['C1'] = 'Note'
        
        # Formatta le intestazioni
        for col in ['A', 'B', 'C']:
            cell = paces_sheet[f'{col}1']
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Imposta larghezze colonne
        paces_sheet.column_dimensions['A'].width = 15
        paces_sheet.column_dimensions['B'].width = 15
        paces_sheet.column_dimensions['C'].width = 30
    
    # Contatori per le righe
    row = 2
    
    # Aggiungi sezioni solo se ci sono valori per esse
    
    # 1. Ritmi per la corsa
    if paces:
        # Aggiungi intestazione sezione Running
        paces_sheet.merge_cells(f'A{row}:C{row}')
        paces_sheet[f'A{row}'] = 'RITMI PER LA CORSA (min/km)'
        paces_sheet[f'A{row}'].font = subheader_font
        paces_sheet[f'A{row}'].fill = running_fill
        paces_sheet[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
        row += 1
        
        # Aggiungi i ritmi dalla configurazione
        for name, value in paces.items():
            paces_sheet[f'A{row}'] = name
            
            # Funzione migliorata per normalizzare i valori ritmo
            normalized_value = format_pace_for_excel(value)
            
            # IMPORTANTE: Speciale gestione per Z3 e altri casi problematici
            if name == 'Z3' and normalized_value == '0:06':
                normalized_value = '6:00'
                print(f"Corretto il valore di Z3 da 0:06 a 6:00")
            
            # Assicurati che i valori che iniziano con 0: vengano convertiti in minuti:00
            if normalized_value and isinstance(normalized_value, str) and re.match(r'^0:\d{2}$', normalized_value):
                minutes = int(normalized_value.split(':')[1])
                normalized_value = f"{minutes}:00"
                print(f"Convertito {value} in {normalized_value}")
            
            # Imposta il valore e applica il formato corretto
            cell = paces_sheet[f'B{row}']
            cell.value = normalized_value
            cell.data_type = TYPE_STRING  # Forza formato testo
            
            # Descrizione di default in base al nome
            description = ""
            if name.startswith('Z'):
                zone_match = re.match(r'Z(\d)', name)
                if zone_match:
                    zone_num = int(zone_match.group(1))
                    if zone_num == 1:
                        description = "Ritmo molto facile (zona 1)"
                    elif zone_num == 2:
                        description = "Ritmo facile (zona 2)"
                    elif zone_num == 3:
                        description = "Ritmo moderato (zona 3)"
                    elif zone_num == 4:
                        description = "Ritmo duro (zona 4)"
                    elif zone_num == 5:
                        description = "Ritmo molto duro (zona 5)"
            elif name == 'recovery':
                description = "Ritmo di recupero"
            elif name == 'threshold':
                description = "Ritmo soglia"
            elif name == 'marathon':
                description = "Ritmo maratona"
            elif name == 'race_pace':
                description = "Ritmo gara"
            
            paces_sheet[f'C{row}'] = description
            
            # Applica bordi e formattazione a tutte le celle della riga
            for col in ['A', 'B', 'C']:
                cell = paces_sheet[f'{col}{row}']
                cell.border = thin_border
                cell.alignment = wrapped_alignment
                
                # Evidenzia le righe in base al tipo di sport attivo
                if sport_type == "running":
                    cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            
            row += 1
        
        # Aggiungi una riga vuota dopo la sezione
        row += 1
    
    # 2. Potenza per il ciclismo
    if power_values:
        # Aggiungi intestazione sezione Power
        paces_sheet.merge_cells(f'A{row}:C{row}')
        paces_sheet[f'A{row}'] = 'POTENZA PER IL CICLISMO (Watt)'
        paces_sheet[f'A{row}'].font = subheader_font
        paces_sheet[f'A{row}'].fill = cycling_fill
        paces_sheet[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
        row += 1
        
        # Aggiungi i valori di potenza dalla configurazione
        for name, value in power_values.items():
            paces_sheet[f'A{row}'] = name
            
            # Imposta il valore come testo
            cell = paces_sheet[f'B{row}']
            cell.value = value
            cell.data_type = TYPE_STRING  # Forza formato testo
            
            # Descrizione di default in base al nome
            description = ""
            if name == 'ftp':
                description = "Functional Threshold Power (W)"
            elif name.startswith('Z'):
                zone_match = re.match(r'Z(\d)', name)
                if zone_match:
                    zone_num = int(zone_match.group(1))
                    if zone_num == 1:
                        description = "Recupero attivo (55-70% FTP)"
                    elif zone_num == 2:
                        description = "Endurance (70-86% FTP)"
                    elif zone_num == 3:
                        description = "Tempo/Soglia (86-100% FTP)"
                    elif zone_num == 4:
                        description = "VO2max (100-120% FTP)"
                    elif zone_num == 5:
                        description = "Capacità anaerobica (120-150% FTP)"
                    elif zone_num == 6:
                        description = "Potenza neuromuscolare (>150% FTP)"
            elif name == 'recovery':
                description = "Recupero (<55% FTP)"
            elif name == 'threshold':
                description = "Soglia (94-106% FTP)"
            elif name == 'sweet_spot':
                description = "Sweet Spot (88-94% FTP)"
            
            paces_sheet[f'C{row}'] = description
            
            # Applica bordi e formattazione a tutte le celle della riga
            for col in ['A', 'B', 'C']:
                cell = paces_sheet[f'{col}{row}']
                cell.border = thin_border
                cell.alignment = wrapped_alignment
                
                # Evidenzia le righe in base al tipo di sport attivo
                if sport_type == "cycling":
                    cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            
            row += 1
        
        # Aggiungi una riga vuota dopo la sezione
        row += 1
    
    # 3. Passi vasca per il nuoto
    if swim_paces:
        # Aggiungi intestazione sezione Swimming
        paces_sheet.merge_cells(f'A{row}:C{row}')
        paces_sheet[f'A{row}'] = 'PASSI VASCA PER IL NUOTO (min/100m)'
        paces_sheet[f'A{row}'].font = subheader_font
        paces_sheet[f'A{row}'].fill = swimming_fill
        paces_sheet[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
        row += 1
        
        # Aggiungi i passi vasca dalla configurazione
        for name, value in swim_paces.items():
            paces_sheet[f'A{row}'] = name
            
            # Imposta il valore e applica il formato corretto
            cell = paces_sheet[f'B{row}']
            cell.value = format_pace_for_excel(value)
            cell.data_type = TYPE_STRING  # Forza formato testo
            
            # Descrizione di default in base al nome
            description = ""
            if name.startswith('Z'):
                zone_match = re.match(r'Z(\d)', name)
                if zone_match:
                    zone_num = int(zone_match.group(1))
                    if zone_num == 1:
                        description = "Passo molto facile (zona 1)"
                    elif zone_num == 2:
                        description = "Passo facile (zona 2)"
                    elif zone_num == 3:
                        description = "Passo moderato (zona 3)"
                    elif zone_num == 4:
                        description = "Passo duro (zona 4)"
                    elif zone_num == 5:
                        description = "Passo molto duro (zona 5)"
            elif name == 'recovery':
                description = "Passo di recupero"
            elif name == 'threshold':
                description = "Passo soglia"
            elif name == 'sprint':
                description = "Passo sprint"
            
            paces_sheet[f'C{row}'] = description
            
            # Applica bordi e formattazione a tutte le celle della riga
            for col in ['A', 'B', 'C']:
                cell = paces_sheet[f'{col}{row}']
                cell.border = thin_border
                cell.alignment = wrapped_alignment
                
                # Evidenzia le righe in base al tipo di sport attivo
                if sport_type == "swimming":
                    cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            
            row += 1
    
    # Aggiungi una nota informativa alla fine
    row += 1
    paces_sheet.merge_cells(f'A{row}:C{row}')
    if sport_type == "running":
        paces_sheet[f'A{row}'] = '* Il tipo di sport attivo è CORSA. Le zone Z1-Z5 si riferiscono ai ritmi in min/km.'
    elif sport_type == "cycling":
        paces_sheet[f'A{row}'] = '* Il tipo di sport attivo è CICLISMO. Per la potenza, usa @pwr prima della zona (es. @pwr Z3).'
    elif sport_type == "swimming":
        paces_sheet[f'A{row}'] = '* Il tipo di sport attivo è NUOTO. Le zone Z1-Z5 si riferiscono ai passi vasca in min/100m.'
    paces_sheet[f'A{row}'].font = Font(italic=True)
    paces_sheet[f'A{row}'].alignment = wrapped_alignment




def format_steps_for_excel(steps, sport_type="running"):
    """
    Formatta i passi per il foglio Excel con la corretta indentazione.
    
    Args:
        steps: Lista di passi dell'allenamento
        sport_type: Tipo di sport ('running', 'cycling' o 'swimming')
        
    Returns:
        Testo formattato dei passi
    """
    formatted_steps = []
    
    for step in steps:
        # Check if this is a metadata step (like sport_type or date)
        if isinstance(step, dict) and len(step) == 1:
            key = list(step.keys())[0]
            if key in ['sport_type', 'date']:
                continue  # Skip metadata steps
        
        # Handle repeat steps
        if isinstance(step, dict) and 'repeat' in step and 'steps' in step:
            # Format the repeat step
            iterations = step['repeat']
            substeps = step['steps']
            
            formatted_steps.append(f"repeat {iterations}:")
            
            # Format the substeps with indentation
            for substep in substeps:
                if isinstance(substep, dict) and len(substep) == 1:
                    # Get the step type and detail
                    substep_type = list(substep.keys())[0]
                    substep_detail = substep[substep_type]
                    
                    # Handle different formats based on sport type
                    if sport_type == "cycling":
                        # For cycling, make sure we have @pwr, @spd, or @hr
                        if '@' in substep_detail and '@pwr' not in substep_detail and '@spd' not in substep_detail and '@hr' not in substep_detail:
                            # Default to @pwr for cycling
                            substep_detail = substep_detail.replace('@', '@pwr ')
                    elif sport_type == "running" or sport_type == "swimming":
                        # For running and swimming, convert @spd to @ and @pwr to @
                        if '@spd' in substep_detail:
                            substep_detail = substep_detail.replace('@spd ', '@')
                        if '@pwr' in substep_detail:
                            substep_detail = substep_detail.replace('@pwr ', '@')
                    
                    # Add indentation with two spaces for substeps
                    formatted_steps.append(f"  {substep_type}: {substep_detail}")
        
        # Handle regular steps (not repeats)
        elif isinstance(step, dict) and len(step) == 1:
            step_type = list(step.keys())[0]
            step_detail = step[step_type]
            
            # Handle different formats based on sport type
            if sport_type == "cycling":
                # For cycling, make sure we have @pwr, @spd, or @hr
                if '@' in step_detail and '@pwr' not in step_detail and '@spd' not in step_detail and '@hr' not in step_detail:
                    # Default to @pwr for cycling
                    step_detail = step_detail.replace('@', '@pwr ')
            elif sport_type == "running" or sport_type == "swimming":
                # For running and swimming, convert @spd to @ and @pwr to @
                if '@spd' in step_detail:
                    step_detail = step_detail.replace('@spd ', '@')
                if '@pwr' in step_detail:
                    step_detail = step_detail.replace('@pwr ', '@')
            
            formatted_steps.append(f"{step_type}: {step_detail}")
    
    return "\n".join(formatted_steps)


def create_unified_examples_sheet(workbook):
    """
    Crea un foglio Examples unificato che contiene esempi per corsa, ciclismo e nuoto.
    
    Args:
        workbook: Workbook di openpyxl
        
    Returns:
        Il foglio Examples creato
    """
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    # Crea o ottieni il foglio Examples
    if 'Examples' in workbook.sheetnames:
        examples_sheet = workbook['Examples']
        # Pulisci il foglio esistente
        for row in range(examples_sheet.max_row, 0, -1):
            examples_sheet.delete_rows(row)
    else:
        examples_sheet = workbook.create_sheet(title='Examples')
    
    # Definisci stili
    header_font = Font(bold=True)
    subheader_font = Font(bold=True, size=12)
    wrapped_alignment = Alignment(wrap_text=True, vertical='top')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    running_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Verde chiaro
    cycling_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")  # Azzurro chiaro
    swimming_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # Arancione chiaro
    alternate_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    
    # Imposta le intestazioni di colonna
    examples_sheet['A1'] = 'Tipo di Esempio'
    examples_sheet['B1'] = 'Descrizione'
    examples_sheet['C1'] = 'Passi (Steps)'
    
    # Formatta le intestazioni
    for col in ['A', 'B', 'C']:
        cell = examples_sheet[f'{col}1']
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Imposta larghezze colonne
    examples_sheet.column_dimensions['A'].width = 20
    examples_sheet.column_dimensions['B'].width = 40
    examples_sheet.column_dimensions['C'].width = 60
    
    # Riga 2: Nota informativa
    examples_sheet.merge_cells('A2:C2')
    examples_sheet['A2'] = '# ESEMPI DI SINTASSI - questo foglio è solo per scopo informativo e non viene importato'
    examples_sheet['A2'].font = Font(italic=True)
    examples_sheet['A2'].alignment = wrapped_alignment
    
    # Sezione Running
    row = 3
    examples_sheet.merge_cells(f'A{row}:C{row}')
    examples_sheet[f'A{row}'] = 'ESEMPI PER LA CORSA (RUNNING)'
    examples_sheet[f'A{row}'].font = subheader_font
    examples_sheet[f'A{row}'].fill = running_fill
    examples_sheet[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    # Esempi per la corsa
    running_examples = [
        # Distanza
        ('Distanza', 'Esempio di allenamento basato su distanza', 
         "warmup: 2km @Z1_HR\ninterval: 5km @Z3\ncooldown: 1km @Z1_HR"),
        
        # Tempo
        ('Tempo', 'Esempio di allenamento basato su tempo', 
         "warmup: 10min @Z1_HR\ninterval: 30min @Z2\ncooldown: 5min @Z1_HR"),
        
        # Ripetute semplici
        ('Ripetute semplici', 'Esempio di allenamento con ripetute', 
         "warmup: 10min @Z1_HR\nrepeat 5:\n  interval: 1km @Z4\n  recovery: 2min @Z1_HR\ncooldown: 10min @Z1_HR"),
        
        # Con descrizioni
        ('Con descrizioni', 'Esempio con descrizioni per ogni passo', 
         "warmup: 10min @Z1_HR -- Inizia lentamente\ninterval: 20min @Z3 -- Mantieni ritmo costante\ncooldown: 5min @Z1_HR -- Rallenta gradualmente"),
        
        # Pulsante lap
        ('Pulsante lap', 'Esempio con pulsante lap', 
         "warmup: 10min @Z1_HR\nrest: lap-button @Z1_HR -- Premi lap quando sei pronto\ninterval: 5km @Z3\ncooldown: 5min @Z1_HR"),
        
        # Zone personalizzate
        ('Zone personalizzate', 'Esempio con zone personalizzate', 
         "warmup: 10min @Z1_HR\ninterval: 20min @marathon\ninterval: 10min @threshold\ncooldown: 5min @Z1_HR")
    ]
    
    # Aggiungi esempi per la corsa
    for i, (tipo, descrizione, passi) in enumerate(running_examples):
        examples_sheet[f'A{row}'] = tipo
        examples_sheet[f'B{row}'] = descrizione
        examples_sheet[f'C{row}'] = passi
        
        # Applica bordi e formattazione a tutte le celle della riga
        for col in ['A', 'B', 'C']:
            cell = examples_sheet[f'{col}{row}']
            cell.border = thin_border
            cell.alignment = wrapped_alignment
            
            # Applica un colore di sfondo alternato per migliorare la leggibilità
            if i % 2 == 0:
                cell.fill = alternate_fill
        
        row += 1
    
    # Aggiungi una riga vuota tra le sezioni
    row += 1
    
    # Sezione Cycling
    examples_sheet.merge_cells(f'A{row}:C{row}')
    examples_sheet[f'A{row}'] = 'ESEMPI PER IL CICLISMO (CYCLING)'
    examples_sheet[f'A{row}'].font = subheader_font
    examples_sheet[f'A{row}'].fill = cycling_fill
    examples_sheet[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    # Esempi per il ciclismo
    cycling_examples = [
        # Potenza - Zone
        ('Potenza (Zone)', 'Allenamento basato su zone di potenza', 
         "warmup: 15min @hr Z1_HR\ninterval: 40min @pwr Z3\ncooldown: 10min @hr Z1_HR"),
        
        # Potenza - Percentuale FTP
        ('Potenza (% FTP)', 'Allenamento basato su percentuali FTP', 
         "warmup: 15min @hr Z1_HR\ninterval: 20min @pwr 90%\ncooldown: 10min @hr Z1_HR"),
        
        # Potenza - Sweet Spot
        ('Sweet Spot', 'Allenamento Sweet Spot (88-94% FTP)', 
         "warmup: 15min @hr Z1_HR\nrepeat 3:\n  interval: 12min @pwr sweet_spot\n  recovery: 3min @hr Z1_HR\ncooldown: 10min @hr Z1_HR"),
        
        # Potenza - Ripetute
        ('Ripetute potenza', 'Ripetute ad alta intensità', 
         "warmup: 15min @hr Z1_HR\nrepeat 5:\n  interval: 3min @pwr Z5\n  recovery: 3min @hr Z1_HR\ncooldown: 10min @hr Z1_HR"),
        
        # Potenza - VO2max
        ('VO2max', 'Allenamento VO2max', 
         "warmup: 15min @hr Z1_HR\nrepeat 5:\n  interval: 3min @pwr 110-120%\n  recovery: 3min @hr Z1_HR\ncooldown: 10min @hr Z1_HR"),
        
        # Potenza - Neuromuscolare
        ('Neuromuscolare', 'Allenamento potenza neuromuscolare', 
         "warmup: 15min @hr Z1_HR\nrepeat 10:\n  interval: 30sec @pwr Z6\n  recovery: 4min30sec @hr Z1_HR\ncooldown: 10min @hr Z1_HR")
    ]
    
    # Aggiungi esempi per il ciclismo
    for i, (tipo, descrizione, passi) in enumerate(cycling_examples):
        examples_sheet[f'A{row}'] = tipo
        examples_sheet[f'B{row}'] = descrizione
        examples_sheet[f'C{row}'] = passi
        
        # Applica bordi e formattazione a tutte le celle della riga
        for col in ['A', 'B', 'C']:
            cell = examples_sheet[f'{col}{row}']
            cell.border = thin_border
            cell.alignment = wrapped_alignment
            
            # Applica un colore di sfondo alternato per migliorare la leggibilità
            if i % 2 == 0:
                cell.fill = alternate_fill
        
        row += 1
    
    # Aggiungi una riga vuota tra le sezioni
    row += 1
    
    # Sezione Swimming (NUOVA)
    examples_sheet.merge_cells(f'A{row}:C{row}')
    examples_sheet[f'A{row}'] = 'ESEMPI PER IL NUOTO (SWIMMING)'
    examples_sheet[f'A{row}'].font = subheader_font
    examples_sheet[f'A{row}'].fill = swimming_fill
    examples_sheet[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    # Esempi per il nuoto
    swimming_examples = [
        # Distanza
        ('Distanza', 'Esempio di allenamento basato su distanza', 
         "warmup: 200m @Z1_HR\ninterval: 800m @Z3\ncooldown: 100m @Z1_HR"),
        
        # Tempo
        ('Tempo', 'Esempio di allenamento basato su tempo', 
         "warmup: 5min @Z1_HR\ninterval: 20min @Z2\ncooldown: 5min @Z1_HR"),
        
        # Ripetute semplici
        ('Ripetute semplici', 'Esempio di allenamento con ripetute', 
         "warmup: 200m @Z1_HR\nrepeat 5:\n  interval: 100m @Z4\n  recovery: 30s @Z1_HR\ncooldown: 100m @Z1_HR"),
        
        # Ripetute con tecniche diverse
        ('Tecniche diverse', 'Esempio di allenamento con diverse tecniche di nuoto', 
         "warmup: 200m @Z1_HR\nrepeat 4:\n  interval: 50m @Z3 -- Stile libero\n  interval: 50m @Z2 -- Dorso\n  recovery: 20s @Z1_HR\ncooldown: 100m @Z1_HR"),
        
        # Con descrizioni
        ('Con descrizioni', 'Esempio con descrizioni per ogni passo', 
         "warmup: 200m @Z1_HR -- Stile libero lento\ninterval: 400m @Z3 -- Alternare stile ogni 100m\ncooldown: 100m @Z1_HR -- Dorso rilassato"),
        
        # Zone personalizzate
        ('Zone personalizzate', 'Esempio con zone personalizzate', 
         "warmup: 200m @Z1_HR\ninterval: 400m @threshold\nrepeat 4:\n  interval: 50m @sprint\n  recovery: 50m @recovery\ncooldown: 100m @Z1_HR")
    ]
    
    # Aggiungi esempi per il nuoto
    for i, (tipo, descrizione, passi) in enumerate(swimming_examples):
        examples_sheet[f'A{row}'] = tipo
        examples_sheet[f'B{row}'] = descrizione
        examples_sheet[f'C{row}'] = passi
        
        # Applica bordi e formattazione a tutte le celle della riga
        for col in ['A', 'B', 'C']:
            cell = examples_sheet[f'{col}{row}']
            cell.border = thin_border
            cell.alignment = wrapped_alignment
            
            # Applica un colore di sfondo alternato per migliorare la leggibilità
            if i % 2 == 0:
                cell.fill = alternate_fill
        
        row += 1
    
    # Aggiungi una riga vuota prima delle note sulla sintassi
    row += 1
    
    # Aggiungi note sulla sintassi generali
    examples_sheet.merge_cells(f'A{row}:C{row}')
    examples_sheet[f'A{row}'] = '# SINTASSI SUPPORTATA NEGLI STEP'
    examples_sheet[f'A{row}'].font = Font(italic=True, bold=True)
    examples_sheet[f'A{row}'].alignment = wrapped_alignment
    row += 1
    
    # Regole di sintassi comuni
    syntax_rules = [
        "Tipi di passo: warmup, interval, recovery, cooldown, rest, repeat, other",
        "Durata: tempo (s, min, h) o distanza (m, km)",
        "Per la corsa: usa @ per il ritmo (es. @Z2) e @hr per la frequenza cardiaca (es. @hr Z2_HR)",
        "Per il ciclismo: usa @pwr per la potenza (es. @pwr Z3 o @pwr 90%), @spd per la velocità (es. @spd Z3) e @hr per FC (es. @hr Z1_HR)",
        "Zone ritmo: Z1-Z5 o qualsiasi zona definita nella sezione Ritmi",
        "Zone potenza: Z1-Z6, percentuali FTP (es. 90%) o zone come sweet_spot, threshold",
        "Zone freq. cardiaca: Z1_HR-Z5_HR o qualsiasi zona definita nel foglio HeartRates",
        "Repeat: repeat N: seguito da step indentati con 2 spazi",
        "Descrizioni opzionali: aggiungi -- seguito dalla descrizione alla fine del passo"
    ]
    
    for rule in syntax_rules:
        examples_sheet.merge_cells(f'A{row}:C{row}')
        examples_sheet[f'A{row}'] = rule
        examples_sheet[f'A{row}'].font = Font(italic=True)
        examples_sheet[f'A{row}'].alignment = wrapped_alignment
        row += 1
    
    return examples_sheet


def update_workouts_sheet(sheet, yaml_data):
    """
    Aggiorna il foglio Workouts con i dati dagli allenamenti YAML.
    
    Args:
        sheet: Foglio Excel Workouts
        yaml_data: Dizionario con i dati YAML
    """
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    # Ottieni lo stile per le intestazioni
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    
    # Definisci un thin border
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Modifica le intestazioni per includere il tipo di sport
    sheet['A2'] = 'Week'
    sheet['B2'] = 'Date'
    sheet['C2'] = 'Session'
    sheet['D2'] = 'Sport'  # Nuova colonna
    sheet['E2'] = 'Description'
    sheet['F2'] = 'Steps'
    
    # Formatta le intestazioni
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        sheet[f'{col}2'].font = Font(bold=True)
        sheet[f'{col}2'].fill = header_fill
    
    # Mantieni le prime due righe (intestazione e atleta)
    for row in range(sheet.max_row, 2, -1):
        sheet.delete_rows(row)
    
    # Ottieni la lista di allenamenti (escluso 'config')
    workouts = []
    for name, steps in yaml_data.items():
        if name != 'config' and isinstance(steps, list):
            # Estrai informazioni dall'allenamento
            match = re.match(r'W(\d+)S(\d+)\s+(.*)', name)
            if match:
                week = int(match.group(1))
                session = int(match.group(2))
                description = match.group(3)
                
                # Estrai la data e il tipo di sport se presenti
                workout_date = ""
                sport_type = "running"  # Default
                
                # Filtra i passi effettivi (escludendo metadati)
                actual_steps = []
                for step in steps:
                    if isinstance(step, dict):
                        if 'sport_type' in step:
                            sport_type = step['sport_type']
                        elif 'date' in step:
                            workout_date = step['date']
                        else:
                            actual_steps.append(step)
                
                # Converti i passi in formato leggibile
                steps_text = format_steps_for_excel(actual_steps, sport_type)
                
                workouts.append((week, session, workout_date, sport_type, description, steps_text))
    
    # Ordina gli allenamenti per settimana e sessione
    workouts.sort(key=lambda x: (x[0], x[1]))
    
    # Definisci i colori per le settimane e gli sport
    week_colors = [
        "FFF2CC",  # Light yellow
        "DAEEF3",  # Light blue
        "E2EFDA",  # Light green
        "FCE4D6",  # Light orange
        "EAD1DC",  # Light pink
        "D9D9D9",  # Light gray
    ]
    
    sport_icons = {
        "running": "🏃",
        "cycling": "🚴",
        "swimming": "🏊",
    }
    
    # Aggiungi gli allenamenti al foglio
    current_week = None
    current_color_index = 0
    
    row = 3  # Prima riga di dati (dopo intestazione e atleta)
    for week, session, workout_date, sport_type, description, steps_text in workouts:
        # Se la settimana cambia, cambia il colore
        if week != current_week:
            current_week = week
            current_color_index = (current_color_index + 1) % len(week_colors)
            
        # Colore di sfondo per la riga corrente
        row_fill = PatternFill(start_color=week_colors[current_color_index], 
                              end_color=week_colors[current_color_index], 
                              fill_type="solid")
        
        # Formatta il tipo di sport
        sport_display = sport_type.capitalize()
        if sport_type in sport_icons:
            sport_display = f"{sport_icons[sport_type]} {sport_display}"
        
        # Assegna valori alle celle
        sheet[f'A{row}'] = week
        sheet[f'B{row}'] = workout_date
        sheet[f'C{row}'] = session
        sheet[f'D{row}'] = sport_display  # Nuova colonna con il tipo di sport
        sheet[f'E{row}'] = description
        sheet[f'F{row}'] = steps_text
        
        # Applica colore di sfondo a tutte le celle della riga
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:  # Aggiunto 'D'
            cell = sheet[f'{col}{row}']
            cell.fill = row_fill
            cell.border = thin_border
            
            # Imposta testo a capo e allineamento
            cell.alignment = Alignment(wrapText=True, vertical='top')
        
        # Calcola altezza appropriata della riga in base al contenuto
        num_lines = 1 + steps_text.count('\n') + steps_text.count(';')
        row_height = max(20, 15 * num_lines)
        sheet.row_dimensions[row].height = row_height
        
        row += 1
    
    # Imposta larghezze colonne
    sheet.column_dimensions['A'].width = 10   # Week
    sheet.column_dimensions['B'].width = 15   # Date
    sheet.column_dimensions['C'].width = 10   # Session
    sheet.column_dimensions['D'].width = 15   # Sport
    sheet.column_dimensions['E'].width = 25   # Description
    sheet.column_dimensions['F'].width = 60   # Steps




def format_steps_for_excel(steps, sport_type="running"):
    """
    Formatta i passi per il foglio Excel con la corretta indentazione.
    
    Args:
        steps: Lista di passi dell'allenamento
        sport_type: Tipo di sport ('running', 'cycling' o 'swimming')
        
    Returns:
        Testo formattato dei passi
    """
    formatted_steps = []
    
    for step in steps:
        if 'repeat' in step and 'steps' in step:
            # Passo di tipo repeat
            iterations = step['repeat']
            substeps = step['steps']
            
            # Formatta il passo di repeat
            formatted_steps.append(f"repeat {iterations}:")
            
            # Formatta i substeps con indentazione
            for substep in substeps:
                if isinstance(substep, dict) and len(substep) == 1:
                    substep_type = list(substep.keys())[0]
                    substep_detail = substep[substep_type]
                    
                    # Gestisci i diversi formati in base al tipo di sport
                    if sport_type == "cycling":
                        # Per il ciclismo, gestisci @pwr, @spd e @hr
                        if '@' in substep_detail:
                            if '@pwr' not in substep_detail and '@spd' not in substep_detail and '@hr' not in substep_detail:
                                # Se c'è @ ma non è specificato il tipo, per ciclismo usa @pwr
                                substep_detail = substep_detail.replace('@', '@pwr ')
                    else:  # running o swimming
                        # Per la corsa e il nuoto, converte eventuali @spd e @pwr in @
                        if '@spd' in substep_detail:
                            substep_detail = substep_detail.replace('@spd ', '@')
                        if '@pwr' in substep_detail:
                            substep_detail = substep_detail.replace('@pwr ', '@')
                    
                    # Usa l'indentazione con due spazi per i substep
                    formatted_steps.append(f"  {substep_type}: {substep_detail}")
        
        elif isinstance(step, dict) and len(step) == 1:
            # Passo normale
            step_type = list(step.keys())[0]
            step_detail = step[step_type]
            
            # Gestisci i diversi formati in base al tipo di sport
            if sport_type == "cycling":
                # Per il ciclismo, gestisci @pwr, @spd e @hr
                if '@' in step_detail:
                    if '@pwr' not in step_detail and '@spd' not in step_detail and '@hr' not in step_detail:
                        # Se c'è @ ma non è specificato il tipo, per ciclismo usa @pwr
                        step_detail = step_detail.replace('@', '@pwr ')
            else:  # running o swimming
                # Per la corsa e il nuoto, converte eventuali @spd e @pwr in @
                if '@spd' in step_detail:
                    step_detail = step_detail.replace('@spd ', '@')
                if '@pwr' in step_detail:
                    step_detail = step_detail.replace('@pwr ', '@')
            
            formatted_steps.append(f"{step_type}: {step_detail}")
    
    return "\n".join(formatted_steps)


def update_heart_rates_sheet(sheet, heart_rates):
    """
    Aggiorna il foglio HeartRates con i dati dalle frequenze cardiache YAML.
    
    Args:
        sheet: Foglio Excel HeartRates
        heart_rates: Dizionario con le frequenze cardiache
    """
    from openpyxl.cell.cell import TYPE_STRING
    
    # Cancella le righe esistenti (tranne l'intestazione)
    for row in range(sheet.max_row, 1, -1):
        sheet.delete_rows(row)
    
    # Aggiungi le nuove righe
    row = 2
    for name, value in heart_rates.items():
        sheet.cell(row=row, column=1).value = name
        
        # Imposta il valore nella colonna B
        cell = sheet.cell(row=row, column=2)
        
        # Se è una stringa che contiene percentuali o trattini, mantienila come testo
        if isinstance(value, str) and ('%' in value or '-' in value):
            cell.value = value
            cell.data_type = TYPE_STRING  # Forza il tipo di dati come stringa
        else:
            # Se è un valore numerico, può rimanere tale
            cell.value = value
        
        row += 1


def update_speeds_sheet(sheet, speeds):
    """
    Aggiorna il foglio Speeds con i dati dalle velocità YAML.
    
    Args:
        sheet: Foglio Excel Speeds
        speeds: Dizionario con le velocità
    """
    # Cancella le righe esistenti (tranne l'intestazione)
    for row in range(sheet.max_row, 1, -1):
        sheet.delete_rows(row)
    
    # Aggiungi le nuove righe
    row = 2
    for name, value in speeds.items():
        sheet.cell(row=row, column=1).value = name
        sheet.cell(row=row, column=2).value = value
        row += 1

def update_paces_sheet(sheet, paces):
    """
    Aggiorna il foglio Paces con i dati dai ritmi YAML.
    
    Args:
        sheet: Foglio Excel Paces
        paces: Dizionario con i ritmi
    """
    # Cancella le righe esistenti (tranne l'intestazione)
    for row in range(sheet.max_row, 1, -1):
        sheet.delete_rows(row)
    
    # Aggiungi le nuove righe
    row = 2
    for name, value in paces.items():
        sheet.cell(row=row, column=1).value = name
        sheet.cell(row=row, column=2).value = value
        row += 1


def update_config_sheet(sheet, config):
    """
    Aggiorna il foglio Config con i dati dalla configurazione YAML.
    Esclude i valori già presenti nel foglio Paces (paces, power_values, swim_paces) 
    e rimuove sport_type poiché non utilizzato altrove.
    
    Args:
        sheet: Foglio Excel Config
        config: Dizionario con la configurazione
    """
    from openpyxl.cell.cell import TYPE_STRING
    
    # Mappa delle righe da aggiornare
    config_rows = {}
    
    # Trova le righe esistenti e mappa le chiavi alle righe
    for row in range(1, sheet.max_row + 1):
        key = sheet.cell(row=row, column=1).value
        if key:
            config_rows[key] = row
    
    # Liste delle chiavi da gestire - RIMOSSO 'sport_type'
    priority_keys = ['name_prefix', 'margins', 'race_day', 'preferred_days', 'athlete_name']
    
    # Chiavi da NON includere nelle righe di Config in quanto già presenti nei loro fogli dedicati
    # o perché non devono essere mostrate nel foglio Config
    excluded_keys = ['paces', 'speeds', 'swim_paces', 'heart_rates', 'power_values', 'sport_type']
    
    # Prima gestisci le chiavi prioritarie
    for key in priority_keys:
        if key in config:
            if key in config_rows:
                row_index = config_rows[key]
            else:
                row_index = sheet.max_row + 1
                sheet.cell(row=row_index, column=1).value = key
            
            if key == 'margins':
                # Aggiorna i margini
                margins = config.get('margins', {})
                # Assicurati che i valori siano stringhe per evitare problemi con i decimali
                if 'faster' in margins and margins['faster'] is not None:
                    value = margins['faster']
                    if isinstance(value, float) and value.is_integer():
                        sheet.cell(row=row_index, column=2).value = str(int(value))
                    else:
                        sheet.cell(row=row_index, column=2).value = str(value)
                
                if 'slower' in margins and margins['slower'] is not None:
                    value = margins['slower']
                    if isinstance(value, float) and value.is_integer():
                        sheet.cell(row=row_index, column=3).value = str(int(value))
                    else:
                        sheet.cell(row=row_index, column=3).value = str(value)
                
                if 'hr_up' in margins and margins['hr_up'] is not None:
                    sheet.cell(row=row_index, column=4).value = margins['hr_up']
                if 'hr_down' in margins and margins['hr_down'] is not None:
                    sheet.cell(row=row_index, column=5).value = margins['hr_down']
                # Aggiungi i margini di potenza in nuove colonne
                if 'power_up' in margins and margins['power_up'] is not None:
                    sheet.cell(row=row_index, column=6).value = margins['power_up']
                if 'power_down' in margins and margins['power_down'] is not None:
                    sheet.cell(row=row_index, column=7).value = margins['power_down']
            elif key == 'preferred_days':
                # Gestisci preferred_days come lista o stringa
                preferred_days = config[key]
                if preferred_days is not None:
                    if isinstance(preferred_days, list):
                        sheet.cell(row=row_index, column=2).value = str(preferred_days)
                    else:
                        sheet.cell(row=row_index, column=2).value = str(preferred_days)
            else:
                # Gestisci altre chiavi normali
                value = config[key]
                if value is not None:
                    # Converti dizionari, liste o altri tipi complessi in stringhe
                    if isinstance(value, (dict, list, tuple, set)):
                        if value:  # Se non è vuoto
                            cell = sheet.cell(row=row_index, column=2)
                            cell.value = str(value)
                            cell.data_type = TYPE_STRING
                        else:
                            sheet.cell(row=row_index, column=2).value = ""  # Dizionario vuoto -> stringa vuota
                    else:
                        sheet.cell(row=row_index, column=2).value = value
    
    # Poi gestisci altre chiavi che potrebbero essere presenti nel config,
    # ma escludiamo le chiavi che non vanno nel foglio Config o sono già gestite nei fogli dedicati
    for key, value in config.items():
        if key not in priority_keys and key not in excluded_keys:
            if key in config_rows:
                row_index = config_rows[key]
            else:
                row_index = sheet.max_row + 1
                sheet.cell(row=row_index, column=1).value = key
            
            # Gestisci tutti gli altri parametri, convertendo tipi complessi in stringhe
            if value is not None:
                if isinstance(value, (dict, list, tuple, set)):
                    if value:  # Se non è vuoto
                        cell = sheet.cell(row=row_index, column=2)
                        cell.value = str(value)
                        cell.data_type = TYPE_STRING
                    else:
                        sheet.cell(row=row_index, column=2).value = ""  # Dizionario vuoto -> stringa vuota
                else:
                    sheet.cell(row=row_index, column=2).value = value


def are_required_columns_present(df, required_cols):
    """
    Check if all required columns are present in the DataFrame.
    
    Args:
        df: DataFrame to check
        required_cols: List of required column names
        
    Returns:
        True if all required columns are present, False otherwise
    """
    return all(col in df.columns for col in required_cols)

def handle_missing_columns(excel_file, required_cols):
    """
    Handle missing columns by trying different header positions or case-insensitive matching.
    
    Args:
        excel_file: Path to the Excel file
        required_cols: List of required column names
        
    Returns:
        DataFrame with corrected columns
        
    Raises:
        ValueError: If required columns cannot be found
    """
    try:
        # Try reading with header in the first row
        df = pd.read_excel(excel_file, sheet_name='Workouts', header=0)
        
        if are_required_columns_present(df, required_cols):
            logging.info("'Workouts' sheet found with headers in the first row.")
            return df
            
        # Try case-insensitive matching
        df_cols_lower = [col.lower() for col in df.columns]
        missing = []
        
        for req_col in required_cols:
            if req_col.lower() not in df_cols_lower:
                missing.append(req_col)
        
        if missing:
            raise ValueError(f"Missing columns in 'Workouts' sheet: {', '.join(missing)}")
        
        # Rename columns for consistency
        rename_map = {}
        for col in df.columns:
            for req_col in required_cols:
                if col.lower() == req_col.lower():
                    rename_map[col] = req_col
        
        df = df.rename(columns=rename_map)
        logging.info("Columns renamed for consistency.")
        return df
        
    except Exception as e:
        raise ValueError(f"Error finding required columns: {str(e)}")

def extract_config(xls, plan):
    """
    Extract configuration information from the Config sheet.
    
    Args:
        xls: ExcelFile object
        plan: Plan dictionary to update
        
    Returns:
        Updated plan dictionary
    """
    if 'Config' in xls.sheet_names:
        try:
            config_df = pd.read_excel(xls, 'Config', header=0)
            
            # Extract name prefix (if present)
            name_prefix_rows = config_df[config_df.iloc[:, 0] == 'name_prefix']
            if not name_prefix_rows.empty:
                # Ensure the prefix ends with a space
                prefix = str(name_prefix_rows.iloc[0, 1]).strip()
                # Add a space at the end if not already there
                if prefix and not prefix.endswith(' '):
                    prefix = prefix + ' '
                plan['config']['name_prefix'] = prefix
            
            # Extract margins (if present)
            margins_rows = config_df[config_df.iloc[:, 0] == 'margins']
            if not margins_rows.empty:
                # Check if there are values for the margins
                if pd.notna(margins_rows.iloc[0, 1]):
                    plan['config']['margins']['faster'] = str(margins_rows.iloc[0, 1]).strip()
                if pd.notna(margins_rows.iloc[0, 2]):
                    plan['config']['margins']['slower'] = str(margins_rows.iloc[0, 2]).strip()
                if pd.notna(margins_rows.iloc[0, 3]):
                    plan['config']['margins']['hr_up'] = int(margins_rows.iloc[0, 3])
                if pd.notna(margins_rows.iloc[0, 4]):
                    plan['config']['margins']['hr_down'] = int(margins_rows.iloc[0, 4])
        except Exception as e:
            logging.warning(f"Error extracting configuration: {str(e)}")
    
    return plan

def extract_paces(xls, plan):
    """
    Extract pace information from the Paces sheet.
    
    Args:
        xls: ExcelFile object
        plan: Plan dictionary to update
        
    Returns:
        Updated plan dictionary
    """
    if 'Paces' in xls.sheet_names:
        try:
            paces_df = pd.read_excel(xls, 'Paces', header=0)
            
            for _, row in paces_df.iterrows():
                # Ensure both name and value are present
                if pd.notna(row.iloc[0]) and pd.notna(row.iloc[1]):
                    name = str(row.iloc[0]).strip()
                    value = str(row.iloc[1]).strip()
                    plan['config']['paces'][name] = value
        except Exception as e:
            logging.warning(f"Error extracting paces: {str(e)}")
    
    return plan

def extract_heart_rates(xls, plan):
    """
    Extract heart rate information from the HeartRates sheet.
    
    Args:
        xls: ExcelFile object
        plan: Plan dictionary to update
        
    Returns:
        Updated plan dictionary
    """
    if 'HeartRates' in xls.sheet_names:
        try:
            hr_df = pd.read_excel(xls, 'HeartRates', header=0)
            
            for _, row in hr_df.iterrows():
                # Ensure both name and value are present
                if pd.notna(row.iloc[0]) and pd.notna(row.iloc[1]):
                    name = str(row.iloc[0]).strip()
                    value = row.iloc[1]
                    
                    # Convert numeric values to integers
                    if isinstance(value, (int, float)) and not pd.isna(value):
                        value = int(value)
                    elif isinstance(value, str) and value.strip().isdigit():
                        value = int(value.strip())
                    else:
                        value = str(value).strip()
                        
                    plan['config']['heart_rates'][name] = value
        except Exception as e:
            logging.warning(f"Error extracting heart rates: {str(e)}")
    
    return plan

def add_comments_to_yaml(yaml_file, descriptions):
    """
    Add comments to the YAML file for workout descriptions.
    
    Args:
        yaml_file: Path to the YAML file
        descriptions: Dictionary with workout names and their descriptions
    """
    try:
        with open(yaml_file, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # Add comments for each workout
        for workout_name, description in descriptions.items():
            # Find the line with the workout name
            pattern = f"^{re.escape(workout_name)}:"
            content = re.sub(pattern, f"{workout_name}: # {description}", content, flags=re.MULTILINE)
        
        # Write the updated content
        with open(yaml_file, 'w', encoding='utf-8') as f:
            f.write(content)
            
        logging.info("Comments added to YAML file")
    except Exception as e:
        logging.warning(f"Error adding comments to YAML: {str(e)}")


def parse_workout_steps(steps_str, workout_name="", sport_type="running"):
    """
    Analizza una stringa contenente i passi dell'allenamento.
    
    Supporta il formato "label: value" o "label: value @ zone" e "repeat n:\\n  step1\\n  step2".
    
    Args:
        steps_str: Stringa con i passi dell'allenamento
        workout_name: Nome dell'allenamento (utile per logging)
        sport_type: Tipo di sport (running, cycling, swimming)
        
    Returns:
        Lista di dizionari, uno per ogni passo
    """
    # Lista che conterrà i passi
    steps = []
    
    # Se la stringa è vuota, restituisci una lista vuota
    if not steps_str or steps_str.strip() == '':
        return steps
    
    # Dividi la stringa in righe e rimuovi spazi iniziali/finali
    lines = [line for line in steps_str.splitlines()]
    
    # Flag per tenere traccia delle ripetizioni
    in_repeat = False
    repeat_count = 0
    repeat_steps = []
    current_repeat = None
    
    # Per ogni riga
    i = 0
    while i < len(lines):
        line = lines[i]
        stripped_line = line.strip()
        
        # Salta righe vuote
        if not stripped_line:
            i += 1
            continue
        
        # Determina l'indentazione
        indent = len(line) - len(line.lstrip())
        
        # Gestisci le ripetizioni
        if stripped_line.lower().startswith('repeat '):
            # Estrai il numero di ripetizioni
            match = re.match(r'repeat\s+(\d+)(?:\s*:|$)', stripped_line)
            if match:
                repeat_count = int(match.group(1))
                
                # Nuova ripetizione, salva quella precedente se esiste
                if in_repeat and current_repeat and repeat_steps:
                    steps.append({'repeat': current_repeat, 'steps': repeat_steps})
                    repeat_steps = []
                
                # Inizia una nuova ripetizione
                in_repeat = True
                current_repeat = repeat_count
                i += 1
                continue
        
        # Se non siamo in una ripetizione e la riga non è indentata, è un passo normale
        if not in_repeat and indent == 0:
            step = parse_step_line(stripped_line, workout_name, sport_type)
            if step:
                steps.append(step)
        
        # Se siamo in una ripetizione e la riga è indentata, è un passo della ripetizione
        elif in_repeat and indent > 0:
            step = parse_step_line(stripped_line, workout_name, sport_type)
            if step:
                repeat_steps.append(step)
        
        # Se siamo in una ripetizione ma la riga non è indentata, è finita la ripetizione
        elif in_repeat and indent == 0:
            # Salva la ripetizione attuale
            if current_repeat and repeat_steps:
                steps.append({'repeat': current_repeat, 'steps': repeat_steps})
                repeat_steps = []
                in_repeat = False
                current_repeat = None
            
            # Processa questa riga come un passo normale
            step = parse_step_line(stripped_line, workout_name, sport_type)
            if step:
                steps.append(step)
        
        i += 1
    
    # Aggiungi l'ultima ripetizione se usciamo dal ciclo ancora in repeat
    if in_repeat and current_repeat and repeat_steps:
        steps.append({'repeat': current_repeat, 'steps': repeat_steps})
    
    return steps


def parse_step_line(line, workout_name="", sport_type="running"):
    """
    Analizza una singola riga di passo.
    
    Args:
        line: Riga da analizzare
        workout_name: Nome dell'allenamento
        sport_type: Tipo di sport
        
    Returns:
        Dizionario con il tipo di passo e i dettagli, o None se non valido
    """
    # Salta righe vuote o commenti
    if not line or line.startswith('#') or line.startswith('//'):
        return None
    
    # Rimuovi eventuali commenti
    if '#' in line:
        line = line.split('#')[0].strip()
    if '//' in line:
        line = line.split('//')[0].strip()
    
    # Cerca di estrarre il tipo di passo e i dettagli
    match = re.match(r'^([\w-]+)\s*:\s*(.+)$', line)
    if match:
        step_type = match.group(1).strip().lower()
        step_detail = match.group(2).strip()
        
        # Verifica se è repeat (questo è solo per compatibilità)
        if step_type == 'repeat' and step_detail.isdigit():
            return {'repeat': int(step_detail), 'steps': []}
        
        # Normalizza i tipi di passi
        if step_type in ['running', 'cycling', 'swimming']:
            # Questo è un tipo di sport, non un passo
            # Ignoralo o aggiungilo come metadato
            step_type = 'interval'
        
        # Sostituzioni per uniformare i tipi di passi
        step_type_map = {
            'run': 'interval',
            'cycle': 'interval',
            'swim': 'interval',
            'bike': 'interval',
            'rest': 'recovery',
            'rec': 'recovery',
            'cool': 'cooldown',
            'warm': 'warmup',
        }
        
        # Applica le sostituzioni se necessario
        if step_type in step_type_map:
            step_type = step_type_map[step_type]
        
        return {step_type: step_detail}
    else:
        # Se non riesci a estrarre il tipo e i dettagli, potrebbe comunque essere un passo valido
        # ma in un formato diverso (come "intervallo 5min Z2")
        words = line.split()
        if len(words) >= 2:
            try:
                step_type = words[0].lower()
                step_detail = ' '.join(words[1:])
                
                # Controlla se il tipo è valido
                valid_types = ['interval', 'recovery', 'warmup', 'cooldown', 'repeat']
                
                if step_type not in valid_types:
                    # Potrebbe essere un tipo diverso, imposta a interval di default
                    step_detail = line
                    step_type = 'interval'
                
                return {step_type: step_detail}
            except:
                # In caso di errore, usa un tipo di passo predefinito
                return {'interval': line}
        else:
            # Se non c'è abbastanza informazione, considera come intervallo generico
            return {'interval': line}


def auto_adjust_column_widths(worksheet):
    """
    Automatically adjust column widths based on content, handling merged cells properly.
    
    Args:
        worksheet: openpyxl worksheet object
    """
    for column_cells in worksheet.columns:
        max_length = 0
        column = None
        
        for cell in column_cells:
            # Skip merged cells which don't have column_letter attribute
            if hasattr(cell, 'column_letter'):
                if column is None:
                    column = cell.column_letter
                
                if cell.value:
                    cell_length = len(str(cell.value))
                    max_length = max(max_length, cell_length)
        
        if column is not None:  # Only adjust if we found a valid column
            adjusted_width = max(max_length + 2, 8)  # Add some extra space
            worksheet.column_dimensions[column].width = min(adjusted_width, 60)  # Limit to 60 to avoid too wide columns



def create_sample_excel(output_file='sample_training_plan.xlsx', sport_type="running"):
    """
    Create a sample Excel file with the expected structure for the training plan.
    Always includes running paces, power values for cycling and swimming paces in a unified Paces sheet.
    Now includes workout examples for all three sports regardless of the primary sport type.
    
    Args:
        output_file: Path for the output Excel file
        sport_type: Type of sport ('running', 'cycling', or 'swimming') - used to determine the primary sport
        
    Returns:
        Path to the created Excel file, or None if there was an error
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.cell.cell import TYPE_STRING
    except ImportError:
        logging.error("ERROR: openpyxl library is not installed.")
        logging.error("Install openpyxl with: pip install openpyxl")
        return None
    
    logging.info(f"Creating sample Excel file: {output_file}")
    
    wb = openpyxl.Workbook()
    
    # Define a thin border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Genera un suffisso casuale per il prefisso del nome
    random_suffix = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
    prefix = f"MYRUN_{random_suffix}_"
    
    # Generate a random athlete name
    athlete_names = [
        "Mario Rossi", "Laura Bianchi", "Andrea Verdi", "Giulia Neri", 
        "Marco Esposito", "Alessia Romano", "Luca Ferrari", "Elena Russo",
        "Giovanni Marino", "Sofia Greco", "Matteo Bruno", "Elisa Ricci"
    ]
    athlete_name = random.choice(athlete_names)
    
    # Config sheet
    config_sheet = wb.active
    config_sheet.title = 'Config'
    
    # Config sheet headers
    config_sheet['A1'] = 'Parameter'
    config_sheet['B1'] = 'Value'
    config_sheet['C1'] = 'Slower'
    config_sheet['D1'] = 'HR Up'
    config_sheet['E1'] = 'HR Down'
    config_sheet['F1'] = 'Power Up'  # Aggiungiamo l'intestazione per Power Up
    config_sheet['G1'] = 'Power Down'  # Aggiungiamo l'intestazione per Power Down
    
    # Config sheet values
    config_sheet['A2'] = 'name_prefix'
    config_sheet['B2'] = prefix
    
    # Aggiungi nome atleta nella configurazione
    config_sheet['A3'] = 'athlete_name'
    config_sheet['B3'] = athlete_name
    
    # Imposta i margini appropriati 
    config_sheet['A4'] = 'margins'
    
    # Imposta valori come stringhe per i campi che potrebbero essere interpretati come orari
    cell_b4 = config_sheet['B4']
    cell_b4.value = '0:03'   # faster in min:sec
    cell_b4.data_type = TYPE_STRING
    
    cell_c4 = config_sheet['C4']
    cell_c4.value = '0:03'   # slower in min:sec
    cell_c4.data_type = TYPE_STRING
    
    config_sheet['D4'] = 5        # hr_up
    config_sheet['E4'] = 5        # hr_down
    config_sheet['F4'] = 10       # power_up in Watt
    config_sheet['G4'] = 10       # power_down in Watt
    
    # Aggiungi race_day (data gara)
    config_sheet['A5'] = 'race_day'
    # Imposta una data di esempio 6 mesi nel futuro
    from datetime import datetime, timedelta
    future_date = datetime.now() + timedelta(days=180)
    config_sheet['B5'] = future_date.strftime("%Y-%m-%d")
    
    # Aggiungi preferred_days (giorni preferiti)
    config_sheet['A6'] = 'preferred_days'
    config_sheet['B6'] = '[1, 3, 5]'  # Martedì, Giovedì, Sabato
    
    # Rimuoviamo sport_type poiché non utilizzato altrove
    
    # Format header
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:  # Includiamo tutte le colonne
        config_sheet[f'{col}1'].font = Font(bold=True)
        config_sheet[f'{col}1'].fill = header_fill
    
    # Crea il foglio Paces unificato con tutti i tipi di ritmi/potenza/passo vasca
    create_unified_paces_sheet(wb, sport_type)
    
    # HeartRates sheet (Z1-Z5 zones)
    hr_sheet = wb.create_sheet(title='HeartRates')
    
    hr_sheet['A1'] = 'Name'
    hr_sheet['B1'] = 'Value'
    
    # Example of using max_hr with percentages
    hr_sheet['A2'] = 'max_hr'
    hr_sheet['B2'] = 180  # Use an integer instead of a string
    
    hr_sheet['A3'] = 'Z1_HR'
    cell_b3 = hr_sheet['B3']
    cell_b3.value = '62-76% max_hr'
    cell_b3.data_type = TYPE_STRING  # Force text format
    
    hr_sheet['A4'] = 'Z2_HR'
    cell_b4 = hr_sheet['B4']
    cell_b4.value = '76-85% max_hr'
    cell_b4.data_type = TYPE_STRING
    
    hr_sheet['A5'] = 'Z3_HR'
    cell_b5 = hr_sheet['B5']
    cell_b5.value = '85-91% max_hr'
    cell_b5.data_type = TYPE_STRING
    
    hr_sheet['A6'] = 'Z4_HR'
    cell_b6 = hr_sheet['B6']
    cell_b6.value = '91-95% max_hr'
    cell_b6.data_type = TYPE_STRING
    
    hr_sheet['A7'] = 'Z5_HR'
    cell_b7 = hr_sheet['B7']
    cell_b7.value = '95-100% max_hr'
    cell_b7.data_type = TYPE_STRING
    
    # Format header
    for col in ['A', 'B']:
        hr_sheet[f'{col}1'].font = Font(bold=True)
        hr_sheet[f'{col}1'].fill = header_fill
    
    # Single Workouts sheet for all workouts
    workouts_sheet = wb.create_sheet(title='Workouts')
    
    # Add athlete name in the merged cell
    workouts_sheet.merge_cells('A1:F1')
    athlete_cell = workouts_sheet['A1']
    athlete_cell.value = f"Atleta: {athlete_name}"
    athlete_cell.alignment = Alignment(horizontal='center', vertical='center')
    athlete_cell.font = Font(size=12, bold=True)
    athlete_cell.border = thin_border

    # Headers in row 2
    workouts_sheet['A2'] = 'Week'
    workouts_sheet['B2'] = 'Date'
    workouts_sheet['C2'] = 'Session'
    workouts_sheet['D2'] = 'Sport'    # Colonna per il tipo di sport
    workouts_sheet['E2'] = 'Description'
    workouts_sheet['F2'] = 'Steps'

    # Format header
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        cell = workouts_sheet[f'{col}2']
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = thin_border  # Add border to all header cells
    
    # Generate sample dates starting from today
    today = datetime.now()
    
    # Define example workouts for all three sports
    running_workouts = [
        # Week, Session, Sport, Description, Steps
        (1, 1, "running", 'Easy run', 'warmup: 10min @ Z1_HR\ninterval: 30min @ Z2\ncooldown: 5min @ Z1_HR'),
        (1, 2, "running", 'Short intervals', 'warmup: 15min @ Z1_HR\nrepeat 5:\n  interval: 400m @ Z5\n  recovery: 2min @ Z1_HR\ncooldown: 10min @ Z1_HR'),
        (1, 3, "running", 'Long slow run', 'warmup: 10min @ Z1_HR\ninterval: 45min @ Z2\ncooldown: 5min @ Z1_HR'),
    ]
    
    cycling_workouts = [
        # Week, Session, Sport, Description, Steps
        (2, 1, "cycling", 'Easy ride', 'warmup: 15min @hr Z1_HR\ninterval: 45min @pwr Z2\ncooldown: 10min @hr Z1_HR'),
        (2, 2, "cycling", 'Sweet Spot', 'warmup: 20min @hr Z1_HR\nrepeat 3:\n  interval: 12min @pwr sweet_spot\n  recovery: 3min @hr Z1_HR\ncooldown: 15min @hr Z1_HR'),
        (2, 3, "cycling", 'Threshold intervals', 'warmup: 15min @hr Z1_HR\nrepeat 4:\n  interval: 8min @pwr threshold\n  recovery: 4min @hr Z1_HR\ncooldown: 10min @hr Z1_HR'),
    ]
    
    swimming_workouts = [
        # Week, Session, Sport, Description, Steps
        (3, 1, "swimming", 'Easy swim', 'warmup: 200m @ Z1_HR\ninterval: 600m @ Z2\ncooldown: 100m @ Z1_HR'),
        (3, 2, "swimming", 'Technique focus', 'warmup: 200m @ Z1_HR\nrepeat 4:\n  interval: 50m @ Z3 -- Tecnica bracciata\n  interval: 50m @ Z3 -- Tecnica gambata\n  recovery: 20s @ Z1_HR\ncooldown: 100m @ Z1_HR'),
        (3, 3, "swimming", 'Sprint intervals', 'warmup: 300m @ Z1_HR\nrepeat 6:\n  interval: 50m @ Z5\n  recovery: 30s @ Z1_HR\ncooldown: 200m @ Z1_HR'),
    ]
    
    # Combine all workouts
    all_workouts = running_workouts + cycling_workouts + swimming_workouts
    
    # Define alternating colors for weeks
    week_colors = [
        "FFF2CC",  # Light yellow
        "DAEEF3",  # Light blue
        "E2EFDA",  # Light green
        "FCE4D6",  # Light orange
        "EAD1DC",  # Light pink
        "D9D9D9",  # Light gray
    ]
    
    # Sport icons
    sport_icons = {
        "running": "🏃",
        "cycling": "🚴",
        "swimming": "🏊",
    }
    
    # Add workouts to the sheet
    current_week = None
    current_color_index = 0
    
    for i, (week, session, sport, description, steps) in enumerate(all_workouts, start=3):
        # If the week changes, change the color
        if week != current_week:
            current_week = week
            current_color_index = (week - 1) % len(week_colors)
        
        # Generate a sample date (today + day offset based on week/session)
        day_offset = (week - 1) * 7 + session  # Simple pattern: week 1/session 1 = day 1, week 1/session 2 = day 2, etc.
        sample_date = today + timedelta(days=day_offset)
        sample_date_str = sample_date.strftime("%Y-%m-%d")
            
        # Background color for the current row
        row_fill = PatternFill(start_color=week_colors[current_color_index], 
                              end_color=week_colors[current_color_index], 
                              fill_type="solid")
        
        # Format sport display
        sport_display = sport.capitalize()
        if sport in sport_icons:
            sport_display = f"{sport_icons[sport]} {sport_display}"
        
        # Assign values to cells
        workouts_sheet[f'A{i}'] = week
        workouts_sheet[f'B{i}'] = sample_date_str  # Add sample date
        workouts_sheet[f'C{i}'] = session
        workouts_sheet[f'D{i}'] = sport_display  # Sport column
        workouts_sheet[f'E{i}'] = description
        
        # Set steps as string type to prevent automatic time format conversion
        cell = workouts_sheet[f'F{i}']
        cell.value = steps
        cell.data_type = TYPE_STRING
        
        # Apply background color and border to all cells in the row
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            cell = workouts_sheet[f'{col}{i}']
            cell.fill = row_fill
            cell.border = thin_border
            
            # Set text wrapping and alignment
            cell.alignment = Alignment(wrapText=True, vertical='top')
        
        # Calculate appropriate row height based on content
        # Count lines of text in steps (both \n and ;)
        num_lines = 1 + steps.count('\n') + steps.count(';')
        
        # Consider indentation for repeats
        if 'repeat' in steps and '\n' in steps:
            # Count indented lines after repeat
            lines_after_repeat = steps.split('repeat')[1].count('\n')
            if lines_after_repeat > 0:
                num_lines += lines_after_repeat - 1  # -1 because the line with 'repeat' is already counted
        
        # Minimum height plus height for each line of text (about 15 points per line)
        row_height = max(20, 15 * num_lines)  # Increased minimum height
        workouts_sheet.row_dimensions[i].height = row_height
    
    # Set column widths
    workouts_sheet.column_dimensions['A'].width = 10  # Week
    workouts_sheet.column_dimensions['B'].width = 15  # Date
    workouts_sheet.column_dimensions['C'].width = 10  # Session
    workouts_sheet.column_dimensions['D'].width = 15  # Sport
    workouts_sheet.column_dimensions['E'].width = 25  # Description
    workouts_sheet.column_dimensions['F'].width = 60  # Steps
    
    # Create Examples sheet with proper content for all sports
    create_unified_examples_sheet(wb)
    
    # Ensure sheets are in the correct order
    sheet_order = ['Config', 'Paces', 'HeartRates', 'Workouts', 'Examples']
    
    # Reorder sheets
    wb._sheets = [wb[sheet_name] for sheet_name in sheet_order if sheet_name in wb.sheetnames]
    
    # Save the file
    wb.save(output_file)
    logging.info(f"Sample Excel file created: {output_file}")
    return output_file


def create_unified_paces_sheet(workbook, sport_type="running"):
    """
    Crea un foglio Paces unificato che contiene sia i ritmi per la corsa,
    le zone di potenza FTP per il ciclismo e i passi vasca per il nuoto.
    
    Args:
        workbook: Workbook di openpyxl
        sport_type: Tipo di sport ('running', 'cycling' o 'swimming')
        
    Returns:
        Il foglio Paces creato
    """
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.cell.cell import TYPE_STRING
    
    # Crea o ottieni il foglio Paces
    if 'Paces' in workbook.sheetnames:
        paces_sheet = workbook['Paces']
        # Pulisci il foglio esistente
        for row in range(paces_sheet.max_row, 0, -1):
            paces_sheet.delete_rows(row)
    else:
        paces_sheet = workbook.create_sheet(title='Paces')
    
    # Definisci stili
    header_font = Font(bold=True)
    subheader_font = Font(bold=True, size=12)
    wrapped_alignment = Alignment(wrap_text=True, vertical='top')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    running_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Verde chiaro
    cycling_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")  # Azzurro chiaro
    power_fill = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")    # Verde oliva
    swimming_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # Arancione chiaro
    
    # Imposta le intestazioni di colonna
    paces_sheet['A1'] = 'Name'
    paces_sheet['B1'] = 'Value'
    paces_sheet['C1'] = 'Note'
    
    # Formatta le intestazioni
    for col in ['A', 'B', 'C']:
        cell = paces_sheet[f'{col}1']
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Imposta larghezze colonne
    paces_sheet.column_dimensions['A'].width = 15
    paces_sheet.column_dimensions['B'].width = 15
    paces_sheet.column_dimensions['C'].width = 30
    
    # Sezione Running
    row = 2
    paces_sheet.merge_cells(f'A{row}:C{row}')
    paces_sheet[f'A{row}'] = 'RITMI PER LA CORSA (min/km)'
    paces_sheet[f'A{row}'].font = subheader_font
    paces_sheet[f'A{row}'].fill = running_fill
    paces_sheet[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    # Ritmi per la corsa
    running_paces = [
        ('Z1', '6:30', 'Ritmo facile (zona 1, recuperativo)'),
        ('Z2', '6:00', 'Ritmo aerobico (zona 2, endurance)'),
        ('Z3', '5:30', 'Ritmo medio (zona 3, soglia aerobica)'),
        ('Z4', '5:00', 'Ritmo soglia (zona 4, soglia anaerobica)'),
        ('Z5', '4:30', 'Ritmo VO2max (zona 5, anaerobico)'),
        ('recovery', '7:00', 'Ritmo recupero (molto lento)'),
        ('threshold', '5:10', 'Ritmo soglia personalizzato'),
        ('marathon', '5:20', 'Ritmo maratona personalizzato'),
        ('race_pace', '5:10', 'Ritmo gara personalizzato'),
    ]
    
    # Aggiungi ritmi per la corsa
    for name, value, note in running_paces:
        paces_sheet[f'A{row}'] = name
        
        # IMPORTANTE: Imposta il valore come stringa e definisci esplicitamente il tipo di dati come stringa (TYPE_STRING)
        # Questo impedisce a Excel di convertire automaticamente in formato orario
        cell = paces_sheet[f'B{row}']
        cell.value = value
        cell.data_type = TYPE_STRING
        
        paces_sheet[f'C{row}'] = note
        
        # Applica bordi e formattazione a tutte le celle della riga
        for col in ['A', 'B', 'C']:
            cell = paces_sheet[f'{col}{row}']
            cell.border = thin_border
            cell.alignment = wrapped_alignment
            
            # Evidenzia le righe in base al tipo di sport attivo
            if sport_type == "running":
                cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        
        row += 1
    
    # Aggiungi una riga vuota tra le sezioni
    row += 1
    
    # NUOVA Sezione - Potenza per il ciclismo
    paces_sheet.merge_cells(f'A{row}:C{row}')
    paces_sheet[f'A{row}'] = 'POTENZA PER IL CICLISMO (Watt)'
    paces_sheet[f'A{row}'].font = subheader_font
    paces_sheet[f'A{row}'].fill = power_fill
    paces_sheet[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    # Valori di potenza per il ciclismo
    cycling_power = [
        ('ftp', '250', 'Functional Threshold Power (W)'),
        ('Z1', '125-175', 'Recupero attivo (55-70% FTP)'),
        ('Z2', '175-215', 'Endurance (70-86% FTP)'),
        ('Z3', '215-250', 'Tempo/Soglia (86-100% FTP)'),
        ('Z4', '250-300', 'VO2max (100-120% FTP)'),
        ('Z5', '300-375', 'Capacità anaerobica (120-150% FTP)'),
        ('Z6', '375+', 'Potenza neuromuscolare (>150% FTP)'),
        ('recovery', '<125', 'Recupero (<55% FTP)'),
        ('threshold', '235-265', 'Soglia (94-106% FTP)'),
        ('sweet_spot', '220-235', 'Sweet Spot (88-94% FTP)'),
    ]
    
    # Aggiungi valori di potenza per il ciclismo
    for name, value, note in cycling_power:
        paces_sheet[f'A{row}'] = name
        
        # Imposta il valore come stringa
        cell = paces_sheet[f'B{row}']
        cell.value = value
        cell.data_type = TYPE_STRING
        
        paces_sheet[f'C{row}'] = note
        
        # Applica bordi e formattazione a tutte le celle della riga
        for col in ['A', 'B', 'C']:
            cell = paces_sheet[f'{col}{row}']
            cell.border = thin_border
            cell.alignment = wrapped_alignment
            
            # Evidenzia le righe in base al tipo di sport attivo
            if sport_type == "cycling":
                cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        
        row += 1
    
    # Aggiungi una riga vuota tra le sezioni
    row += 1
    
    # Sezione Swimming
    paces_sheet.merge_cells(f'A{row}:C{row}')
    paces_sheet[f'A{row}'] = 'PASSI VASCA PER IL NUOTO (min/100m)'
    paces_sheet[f'A{row}'].font = subheader_font
    paces_sheet[f'A{row}'].fill = swimming_fill
    paces_sheet[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    # Passi vasca per il nuoto
    swimming_paces = [
        ('Z1', '2:30', 'Ritmo facile (zona 1)'),
        ('Z2', '2:15', 'Ritmo aerobico (zona 2)'),
        ('Z3', '2:00', 'Ritmo medio (zona 3)'),
        ('Z4', '1:45', 'Ritmo soglia (zona 4)'),
        ('Z5', '1:30', 'Ritmo VO2max (zona 5)'),
        ('recovery', '2:45', 'Ritmo recupero (molto lento)'),
        ('threshold', '1:55', 'Ritmo soglia personalizzato'),
        ('sprint', '1:25', 'Ritmo sprint personalizzato'),
    ]
    
    # Aggiungi passi vasca per il nuoto
    for name, value, note in swimming_paces:
        paces_sheet[f'A{row}'] = name
        
        # Imposta il valore come stringa
        cell = paces_sheet[f'B{row}']
        cell.value = value
        cell.data_type = TYPE_STRING
        
        paces_sheet[f'C{row}'] = note
        
        # Applica bordi e formattazione a tutte le celle della riga
        for col in ['A', 'B', 'C']:
            cell = paces_sheet[f'{col}{row}']
            cell.border = thin_border
            cell.alignment = wrapped_alignment
            
            # Evidenzia le righe in base al tipo di sport attivo
            if sport_type == "swimming":
                cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        
        row += 1
    
    # Aggiungi una nota informativa alla fine
    row += 1
    paces_sheet.merge_cells(f'A{row}:C{row}')
    if sport_type == "running":
        paces_sheet[f'A{row}'] = '* Il tipo di sport attivo è CORSA. Le zone Z1-Z5 si riferiscono ai ritmi in min/km.'
    elif sport_type == "cycling":
        paces_sheet[f'A{row}'] = '* Il tipo di sport attivo è CICLISMO. Per la potenza, usa @pwr prima della zona (es. @pwr Z3).'
    elif sport_type == "swimming":
        paces_sheet[f'A{row}'] = '* Il tipo di sport attivo è NUOTO. Le zone Z1-Z5 si riferiscono ai passi vasca in min/100m.'
    paces_sheet[f'A{row}'].font = Font(italic=True)
    paces_sheet[f'A{row}'].alignment = wrapped_alignment
    
    return paces_sheet



def create_examples_sheet(workbook, sport_type="running"):
    """
    Crea un foglio Examples nel formato corretto in italiano con formattazione appropriata,
    adattato al tipo di sport specificato.
    
    Args:
        workbook: Workbook di openpyxl
        sport_type: Tipo di sport ('running' o 'cycling')
        
    Returns:
        Il foglio Examples creato
    """
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    # Crea o ottieni il foglio Examples
    if 'Examples' in workbook.sheetnames:
        examples_sheet = workbook['Examples']
        # Pulisci il foglio esistente
        for row in range(examples_sheet.max_row, 0, -1):
            examples_sheet.delete_rows(row)
    else:
        examples_sheet = workbook.create_sheet(title='Examples')
    
    # Definisci stili
    header_font = Font(bold=True)
    wrapped_alignment = Alignment(wrap_text=True, vertical='top')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    alternate_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    
    # Imposta le intestazioni di colonna
    examples_sheet['A1'] = 'Tipo di Esempio'
    examples_sheet['B1'] = 'Descrizione'
    examples_sheet['C1'] = 'Passi (Steps)'
    
    # Formatta le intestazioni
    for col in ['A', 'B', 'C']:
        cell = examples_sheet[f'{col}1']
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Imposta larghezze colonne
    examples_sheet.column_dimensions['A'].width = 20
    examples_sheet.column_dimensions['B'].width = 40
    examples_sheet.column_dimensions['C'].width = 60
    
    # Riga 2: Nota informativa
    examples_sheet.merge_cells('A2:C2')
    examples_sheet['A2'] = '# ESEMPI DI SINTASSI - questo foglio è solo per scopo informativo e non viene importato'
    examples_sheet['A2'].font = Font(italic=True)
    examples_sheet['A2'].alignment = wrapped_alignment
    
    # Compila dati di esempio in base al tipo di sport
    examples_data = []
    
    if sport_type == "running":
        # Esempi per la corsa
        examples_data = [
            # Distanza
            ('Distanza', 'Esempio di allenamento basato su distanza', 
             "warmup: 2km @Z1_HR\ninterval: 5km @Z3\ncooldown: 1km @Z1_HR"),
            
            # Tempo
            ('Tempo', 'Esempio di allenamento basato su tempo', 
             "warmup: 10min @Z1_HR\ninterval: 30min @Z2\ncooldown: 5min @Z1_HR"),
            
            # Ripetute semplici
            ('Ripetute semplici', 'Esempio di allenamento con ripetute', 
             "warmup: 10min @Z1_HR\nrepeat 5:\n  interval: 1km @Z4\n  recovery: 2min @Z1_HR\ncooldown: 10min @Z1_HR"),
            
            # Ripetute annidate
            ('Ripetute annidate', 'Esempio di ripetute annidate', 
             "warmup: 10min @Z1_HR\nrepeat 3:\n  interval: 5min @Z3\n  repeat 4:\n    interval: 30s @Z5\n    recovery: 30s @Z1_HR\n  recovery: 3min @Z2_HR\ncooldown: 10min @Z1_HR"),
            
            # Con descrizioni
            ('Con descrizioni', 'Esempio con descrizioni per ogni passo', 
             "warmup: 10min @Z1_HR -- Inizia lentamente\ninterval: 20min @Z3 -- Mantieni ritmo costante\ncooldown: 5min @Z1_HR -- Rallenta gradualmente"),
            
            # Pulsante lap
            ('Pulsante lap', 'Esempio con pulsante lap', 
             "warmup: 10min @Z1_HR\nrest: lap-button @Z1_HR -- Premi lap quando sei pronto\ninterval: 5km @Z3\ncooldown: 5min @Z1_HR"),
            
            # Zone personalizzate
            ('Zone personalizzate', 'Esempio con zone personalizzate', 
             "warmup: 10min @Z1_HR\ninterval: 20min @marathon\ninterval: 10min @threshold\ncooldown: 5min @Z1_HR")
        ]
    else:  # cycling
        # Esempi per il ciclismo
        examples_data = [
            # Distanza
            ('Distanza', 'Esempio di allenamento basato su distanza', 
             "warmup: 5km @hr Z1_HR\ninterval: 20km @spd Z3\ncooldown: 3km @hr Z1_HR"),
            
            # Tempo
            ('Tempo', 'Esempio di allenamento basato su tempo', 
             "warmup: 15min @hr Z1_HR\ninterval: 40min @spd Z2\ncooldown: 5min @hr Z1_HR"),
            
            # Ripetute semplici
            ('Ripetute semplici', 'Esempio di allenamento con ripetute', 
             "warmup: 15min @hr Z1_HR\nrepeat 5:\n  interval: 2min @spd Z4\n  recovery: 2min @hr Z1_HR\ncooldown: 10min @hr Z1_HR"),
            
            # Ripetute annidate
            ('Ripetute annidate', 'Esempio di ripetute annidate', 
             "warmup: 15min @hr Z1_HR\nrepeat 3:\n  interval: 8min @spd Z3\n  repeat 4:\n    interval: 20s @spd Z5\n    recovery: 40s @hr Z1_HR\n  recovery: 3min @hr Z2_HR\ncooldown: 10min @hr Z1_HR"),
            
            # Con descrizioni
            ('Con descrizioni', 'Esempio con descrizioni per ogni passo', 
             "warmup: 15min @hr Z1_HR -- Inizia lentamente\ninterval: 30min @spd Z3 -- Mantieni ritmo costante\ncooldown: 10min @hr Z1_HR -- Rallenta gradualmente"),
            
            # Pulsante lap
            ('Pulsante lap', 'Esempio con pulsante lap', 
             "warmup: 15min @hr Z1_HR\nrest: lap-button @hr Z1_HR -- Premi lap quando sei pronto\ninterval: 20km @spd Z3\ncooldown: 5min @hr Z1_HR"),
            
            # Zone personalizzate
            ('Zone personalizzate', 'Esempio con zone personalizzate', 
             "warmup: 15min @hr Z1_HR\ninterval: 30min @spd 28.0\ninterval: 20min @spd threshold\ncooldown: 10min @hr Z1_HR")
        ]
    
    # Aggiungi esempi alla tabella
    for i, (tipo, descrizione, passi) in enumerate(examples_data, start=3):
        row_num = i
        examples_sheet[f'A{row_num}'] = tipo
        examples_sheet[f'B{row_num}'] = descrizione
        examples_sheet[f'C{row_num}'] = passi
        
        # Applica bordi e formattazione a tutte le celle della riga
        for col in ['A', 'B', 'C']:
            cell = examples_sheet[f'{col}{row_num}']
            cell.border = thin_border
            cell.alignment = wrapped_alignment
            
            # Applica un colore di sfondo alternato per migliorare la leggibilità
            if row_num % 2 == 1:
                cell.fill = alternate_fill
    
    # Aggiungi una riga vuota prima delle note sulla sintassi
    row_num = len(examples_data) + 3
    
    # Aggiungi note sulla sintassi, diverse per ogni sport
    row_num += 1
    examples_sheet.merge_cells(f'A{row_num}:C{row_num}')
    examples_sheet[f'A{row_num}'] = '# SINTASSI SUPPORTATA NEGLI STEP'
    examples_sheet[f'A{row_num}'].font = Font(italic=True)
    examples_sheet[f'A{row_num}'].alignment = wrapped_alignment
    
    # Regole di sintassi specifiche per sport
    if sport_type == "running":
        syntax_rules = [
            "Tipo: durata @zona -- descrizione opzionale",
            "Tipi: warmup, interval, recovery, cooldown, rest, repeat, other",
            "Durata: tempo (s, min, h) o distanza (m, km)",
            "Zone: Z1-Z5 (passo), Z1_HR-Z5_HR (freq. cardiaca), o qualsiasi zona definita nei fogli Paces/HeartRates",
            "Repeat: repeat N: seguito da step indentati con 2 spazi"
        ]
    else:  # cycling
        syntax_rules = [
            "Tipo: durata @hr zona o durata @spd zona -- descrizione opzionale",
            "Tipi: warmup, interval, recovery, cooldown, rest, repeat, other",
            "Durata: tempo (s, min, h) o distanza (m, km)",
            "Zone HR: Z1_HR-Z5_HR (freq. cardiaca) o qualsiasi zona HR definita nel foglio HeartRates",
            "Zone Velocità: Z1-Z5 (km/h) o qualsiasi zona definita nel foglio Speeds",
            "Repeat: repeat N: seguito da step indentati con 2 spazi"
        ]
    
    for i, rule in enumerate(syntax_rules, start=1):
        rule_row = row_num + i
        examples_sheet.merge_cells(f'A{rule_row}:C{rule_row}')
        examples_sheet[f'A{rule_row}'] = rule
        examples_sheet[f'A{rule_row}'].font = Font(italic=True)
        examples_sheet[f'A{rule_row}'].alignment = wrapped_alignment
    
    return examples_sheet



def create_unified_examples_sheet(workbook):
    """
    Crea un foglio Examples unificato che contiene esempi sia per la corsa, il ciclismo e il nuoto.
    
    Args:
        workbook: Workbook di openpyxl
        
    Returns:
        Il foglio Examples creato
    """
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    # Crea o ottieni il foglio Examples
    if 'Examples' in workbook.sheetnames:
        examples_sheet = workbook['Examples']
        # Pulisci il foglio esistente
        for row in range(examples_sheet.max_row, 0, -1):
            examples_sheet.delete_rows(row)
    else:
        examples_sheet = workbook.create_sheet(title='Examples')
    
    # Definisci stili
    header_font = Font(bold=True)
    subheader_font = Font(bold=True, size=12)
    wrapped_alignment = Alignment(wrap_text=True, vertical='top')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    running_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Verde chiaro
    cycling_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")  # Azzurro chiaro
    swimming_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # Arancione chiaro
    alternate_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    
    # Imposta le intestazioni di colonna
    examples_sheet['A1'] = 'Tipo di Esempio'
    examples_sheet['B1'] = 'Descrizione'
    examples_sheet['C1'] = 'Passi (Steps)'
    
    # Formatta le intestazioni
    for col in ['A', 'B', 'C']:
        cell = examples_sheet[f'{col}1']
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Imposta larghezze colonne
    examples_sheet.column_dimensions['A'].width = 20
    examples_sheet.column_dimensions['B'].width = 40
    examples_sheet.column_dimensions['C'].width = 60
    
    # Riga 2: Nota informativa
    examples_sheet.merge_cells('A2:C2')
    examples_sheet['A2'] = '# ESEMPI DI SINTASSI - questo foglio è solo per scopo informativo e non viene importato'
    examples_sheet['A2'].font = Font(italic=True)
    examples_sheet['A2'].alignment = wrapped_alignment
    
    # Sezione Running
    row = 3
    examples_sheet.merge_cells(f'A{row}:C{row}')
    examples_sheet[f'A{row}'] = 'ESEMPI PER LA CORSA (RUNNING)'
    examples_sheet[f'A{row}'].font = subheader_font
    examples_sheet[f'A{row}'].fill = running_fill
    examples_sheet[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    # Esempi per la corsa - ora più dettagliati ed esplicativi
    running_examples = [
        # Durata basata su tempo
        ('Tempo', 'Allenamento continuo basato su tempo', 
         "warmup: 10min @Z1_HR -- Riscaldamento lento\ninterval: 30min @Z2 -- Ritmo aerobico costante\ncooldown: 5min @Z1_HR -- Defaticamento"),
        
        # Durata basata su distanza
        ('Distanza', 'Allenamento continuo basato su distanza', 
         "warmup: 2km @Z1_HR -- Riscaldamento\ninterval: 5km @Z3 -- Ritmo medio\ncooldown: 1km @Z1_HR -- Defaticamento"),
        
        # Ripetute classiche
        ('Ripetute', 'Classiche ripetute con recupero', 
         "warmup: 10min @Z1_HR\nrepeat 5:\n  interval: 400m @Z5 -- Ritmo veloce\n  recovery: 2min @Z1_HR -- Recupero attivo\ncooldown: 10min @Z1_HR"),
        
        # Ripetute con zone personalizzate
        ('Zone personalizzate', 'Utilizzo di zone personalizzate', 
         "warmup: 15min @Z1_HR\ninterval: 15min @marathon -- Ritmo maratona\ninterval: 10min @threshold -- Ritmo soglia\ninterval: 5min @race_pace -- Ritmo gara\ncooldown: 10min @Z1_HR"),
        
        # Ripetute con passo specifico
        ('Passo specifico', 'Utilizzo di un passo specifico invece di una zona', 
         "warmup: 10min @Z1_HR\nrepeat 4:\n  interval: 800m @4:30 -- Ritmo specifico 4:30 min/km\n  recovery: 3min @Z1_HR\ncooldown: 10min @Z1_HR"),
        
        # Pulsante Lap
        ('Pulsante Lap', 'Utilizzo del pulsante Lap per terminare un passo', 
         "warmup: 10min @Z1_HR\nrest: lap-button -- Premi il pulsante Lap quando sei pronto\ninterval: 5km @Z4\ncooldown: 5min @Z1_HR"),
        
        # Allenamento con zone HR
        ('Zone frequenza cardiaca', 'Utilizzo di zone di frequenza cardiaca', 
         "warmup: 10min @hr Z1_HR -- FC bassa\ninterval: 20min @hr Z3_HR -- FC moderata\ninterval: 10min @hr Z4_HR -- FC elevata\ncooldown: 5min @hr Z1_HR -- FC bassa"),
        
        # Ripetute a piramide
        ('Piramide', 'Allenamento a piramide con distanze crescenti e decrescenti', 
         "warmup: 10min @Z1_HR\nrepeat 1:\n  interval: 400m @Z4\n  recovery: 2min @Z1_HR\n  interval: 800m @Z4\n  recovery: 3min @Z1_HR\n  interval: 1200m @Z4\n  recovery: 3min @Z1_HR\n  interval: 800m @Z4\n  recovery: 2min @Z1_HR\n  interval: 400m @Z4\ncooldown: 10min @Z1_HR"),
    ]
    
    # Aggiungi esempi per la corsa
    for i, (tipo, descrizione, passi) in enumerate(running_examples):
        examples_sheet[f'A{row}'] = tipo
        examples_sheet[f'B{row}'] = descrizione
        examples_sheet[f'C{row}'] = passi
        
        # Applica bordi e formattazione a tutte le celle della riga
        for col in ['A', 'B', 'C']:
            cell = examples_sheet[f'{col}{row}']
            cell.border = thin_border
            cell.alignment = wrapped_alignment
            
            # Applica un colore di sfondo alternato per migliorare la leggibilità
            if i % 2 == 0:
                cell.fill = alternate_fill
        
        row += 1
    
    # Aggiungi una riga vuota tra le sezioni
    row += 1
    
    # Sezione Cycling
    examples_sheet.merge_cells(f'A{row}:C{row}')
    examples_sheet[f'A{row}'] = 'ESEMPI PER IL CICLISMO (CYCLING)'
    examples_sheet[f'A{row}'].font = subheader_font
    examples_sheet[f'A{row}'].fill = cycling_fill
    examples_sheet[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    # Esempi per il ciclismo - migliorati con focus sulla potenza e con esempi più pratici
    cycling_examples = [
        # Potenza - Zone FTP
        ('Potenza (Zone FTP)', 'Allenamento con zone di potenza basate sull\'FTP', 
         "warmup: 15min @hr Z1_HR -- FC bassa per riscaldamento\ninterval: 30min @pwr Z3 -- Zona 3 (86-100% FTP)\ncooldown: 10min @hr Z1_HR -- FC bassa per defaticamento"),
        
        # Potenza - Valori FTP specifici
        ('Potenza (% FTP)', 'Allenamento con percentuali specifiche dell\'FTP', 
         "warmup: 15min @hr Z1_HR\ninterval: 20min @pwr 90% -- 90% dell'FTP\ncooldown: 10min @hr Z1_HR"),
        
        # Potenza - Intervallo percentuale
        ('Potenza (range %)', 'Allenamento con intervallo percentuale dell\'FTP', 
         "warmup: 15min @hr Z1_HR\ninterval: 20min @pwr 75-85% -- Tra 75% e 85% dell'FTP\ncooldown: 10min @hr Z1_HR"),
        
        # Potenza - Sweet Spot
        ('Sweet Spot', 'Allenamento "Sweet Spot" (88-94% FTP)', 
         "warmup: 15min @hr Z1_HR\nrepeat 3:\n  interval: 12min @pwr sweet_spot -- 88-94% FTP\n  recovery: 3min @hr Z1_HR\ncooldown: 10min @hr Z1_HR"),
        
        # Potenza - Intervalli
        ('Intervalli di potenza', 'Intervalli ad alta intensità con recupero', 
         "warmup: 15min @hr Z1_HR\nrepeat 5:\n  interval: 3min @pwr Z5 -- Zona 5 (120-150% FTP)\n  recovery: 3min @hr Z1_HR\ncooldown: 10min @hr Z1_HR"),
        
        # Potenza - VO2max
        ('VO2max', 'Intervalli al 110-120% dell\'FTP per sviluppare il VO2max', 
         "warmup: 15min @hr Z1_HR\nrepeat 5:\n  interval: 3min @pwr 110-120% -- Oltre soglia\n  recovery: 3min @hr Z1_HR\ncooldown: 10min @hr Z1_HR"),
        
        # Potenza - Neuromuscolare
        ('Neuromuscolare', 'Sprint brevi ad altissima intensità', 
         "warmup: 15min @hr Z1_HR\nrepeat 10:\n  interval: 30sec @pwr Z6 -- Potenza massimale\n  recovery: 4min30sec @hr Z1_HR\ncooldown: 10min @hr Z1_HR"),
        
        # Potenza - Threshold
        ('Threshold', 'Blocchi di soglia con recupero breve', 
         "warmup: 15min @hr Z1_HR\nrepeat 3:\n  interval: 10min @pwr threshold -- Zona soglia\n  recovery: 2min @hr Z1_HR\ncooldown: 15min @hr Z1_HR"),
    ]
    
    # Aggiungi esempi per il ciclismo
    for i, (tipo, descrizione, passi) in enumerate(cycling_examples):
        examples_sheet[f'A{row}'] = tipo
        examples_sheet[f'B{row}'] = descrizione
        examples_sheet[f'C{row}'] = passi
        
        # Applica bordi e formattazione a tutte le celle della riga
        for col in ['A', 'B', 'C']:
            cell = examples_sheet[f'{col}{row}']
            cell.border = thin_border
            cell.alignment = wrapped_alignment
            
            # Applica un colore di sfondo alternato per migliorare la leggibilità
            if i % 2 == 0:
                cell.fill = alternate_fill
        
        row += 1
    
    # Aggiungi una riga vuota tra le sezioni
    row += 1
    
    # Sezione Swimming (NUOVA)
    examples_sheet.merge_cells(f'A{row}:C{row}')
    examples_sheet[f'A{row}'] = 'ESEMPI PER IL NUOTO (SWIMMING)'
    examples_sheet[f'A{row}'].font = subheader_font
    examples_sheet[f'A{row}'].fill = swimming_fill
    examples_sheet[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    # Esempi per il nuoto - nuova sezione dettagliata
    swimming_examples = [
        # Distanza continua
        ('Distanza continua', 'Nuotata continua di resistenza', 
         "warmup: 200m @Z1_HR -- Riscaldamento lento\ninterval: 1000m @Z2 -- Ritmo costante\ncooldown: 100m @Z1_HR -- Defaticamento"),
        
        # Allenamento a intervalli
        ('Intervalli', 'Intervalli con recupero per il nuoto', 
         "warmup: 200m @Z1_HR\nrepeat 5:\n  interval: 100m @Z4 -- Ritmo veloce\n  recovery: 30s @Z1_HR -- Recupero breve\ncooldown: 100m @Z1_HR"),
        
        # Tecniche di nuoto
        ('Tecniche diverse', 'Allenamento con diverse tecniche di nuoto', 
         "warmup: 200m @Z1_HR -- Stile libero lento\nrepeat 4:\n  interval: 50m @Z3 -- Stile libero\n  interval: 50m @Z2 -- Dorso\n  recovery: 20s @Z1_HR\ncooldown: 100m @Z1_HR -- Nuoto lento a scelta"),
        
        # Sprint 
        ('Sprint', 'Allenamento con sprint brevi e massimali', 
         "warmup: 300m @Z1_HR\nrepeat 8:\n  interval: 25m @sprint -- Sprint massimale\n  recovery: 45s @Z1_HR -- Recupero completo\ncooldown: 200m @Z1_HR"),
        
        # Threshold
        ('Threshold', 'Allenamento alla soglia anaerobica', 
         "warmup: 300m @Z1_HR\nrepeat 3:\n  interval: 200m @threshold -- Ritmo soglia\n  recovery: 45s @Z1_HR\ncooldown: 200m @Z1_HR"),
        
        # Piramide
        ('Piramide', 'Allenamento a piramide con distanze crescenti e decrescenti', 
         "warmup: 200m @Z1_HR\nrepeat 1:\n  interval: 50m @Z4\n  recovery: 20s @Z1_HR\n  interval: 100m @Z4\n  recovery: 30s @Z1_HR\n  interval: 150m @Z4\n  recovery: 40s @Z1_HR\n  interval: 100m @Z4\n  recovery: 30s @Z1_HR\n  interval: 50m @Z4\ncooldown: 100m @Z1_HR"),
        
        # Tecnica con lap-button
        ('Tecnica - Pulsante Lap', 'Esercizi tecnici terminati con pulsante Lap', 
         "warmup: 200m @Z1_HR\nrepeat 5:\n  rest: lap-button -- Premi lap quando sei pronto\n  interval: 50m @Z3 -- Focus sulla tecnica delle bracciate\n  recovery: 15s @Z1_HR\ncooldown: 100m @Z1_HR"),
        
        # Mix di stili
        ('Mix di stili', 'Combinazione di diversi stili di nuoto', 
         "warmup: 200m @Z1_HR -- Stile libero\ninterval: 200m @Z2 -- Dorso\ninterval: 200m @Z2 -- Rana\ninterval: 200m @Z2 -- Stile libero\ncooldown: 100m @Z1_HR -- Stile a scelta")
    ]
    
    # Aggiungi esempi per il nuoto
    for i, (tipo, descrizione, passi) in enumerate(swimming_examples):
        examples_sheet[f'A{row}'] = tipo
        examples_sheet[f'B{row}'] = descrizione
        examples_sheet[f'C{row}'] = passi
        
        # Applica bordi e formattazione a tutte le celle della riga
        for col in ['A', 'B', 'C']:
            cell = examples_sheet[f'{col}{row}']
            cell.border = thin_border
            cell.alignment = wrapped_alignment
            
            # Applica un colore di sfondo alternato per migliorare la leggibilità
            if i % 2 == 0:
                cell.fill = alternate_fill
        
        row += 1
    
    # Aggiungi una riga vuota prima delle note sulla sintassi
    row += 1
    
    # Aggiungi note sulla sintassi generali
    examples_sheet.merge_cells(f'A{row}:C{row}')
    examples_sheet[f'A{row}'] = '# SINTASSI SUPPORTATA NEGLI STEP'
    examples_sheet[f'A{row}'].font = Font(italic=True, bold=True)
    examples_sheet[f'A{row}'].alignment = wrapped_alignment
    row += 1
    
    # Regole di sintassi comuni - Espanse e più dettagliate
    syntax_rules = [
        "Formato generale: tipo_passo: misura [@target] [-- descrizione]",
        "Tipi di passo: warmup (riscaldamento), interval (intervallo), recovery (recupero), cooldown (defaticamento), rest (riposo), repeat (ripetizione), other (altro)",
        "Misura: tempo (10min, 1h, ecc.) o distanza (400m, 5km, ecc.)",
        "Target per corsa: usa @ per il ritmo (es. @Z2, @marathon, @4:30) e @hr per la frequenza cardiaca (es. @hr Z2_HR)",
        "Target per ciclismo: usa @pwr per la potenza (es. @pwr Z3, @pwr 90%, @pwr 220-250) e @hr per la FC (es. @hr Z1_HR)",
        "Target per nuoto: usa @ per il passo vasca (es. @Z2, @threshold, @1:45) e @hr per la FC (es. @hr Z2_HR)",
        "Zone ritmo/velocità: Z1-Z5 o qualsiasi zona definita nel foglio Paces",
        "Zone potenza: Z1-Z6, percentuali come 90% o 75-85%, o zone come sweet_spot, threshold",
        "Zone freq. cardiaca: Z1_HR-Z5_HR o qualsiasi zona definita nel foglio HeartRates",
        "Per ripetizioni: repeat N: seguito da step indentati con 2 spazi",
        "Descrizioni opzionali: aggiungi -- seguito dalla descrizione alla fine del passo"
    ]
    
    for rule in syntax_rules:
        examples_sheet.merge_cells(f'A{row}:C{row}')
        examples_sheet[f'A{row}'] = rule
        examples_sheet[f'A{row}'].font = Font(italic=True)
        examples_sheet[f'A{row}'].alignment = wrapped_alignment
        row += 1
    
    return examples_sheet

def safe_adjust_column_widths(worksheet):
    """
    Automatically adjust column widths based on content, safely handling merged cells.
    
    Args:
        worksheet: openpyxl worksheet object
    """
    for column_cells in worksheet.columns:
        max_length = 0
        column = None
        
        for cell in column_cells:
            # Skip merged cells which don't have column_letter attribute
            if hasattr(cell, 'column_letter'):
                if column is None:
                    column = cell.column_letter
                
                if cell.value:
                    cell_length = len(str(cell.value))
                    max_length = max(max_length, cell_length)
        
        if column is not None:  # Only adjust if we found a valid column
            adjusted_width = max(max_length + 2, 8)  # Add some extra space
            worksheet.column_dimensions[column].width = min(adjusted_width, 60)  # Limit to 60 to avoid too wide columns


def excel_to_yaml(excel_file, output_file=None, sport_type=None):
    """
    Converte un file Excel strutturato in un file YAML compatibile con garmin-planner.
    Include supporto per estrarre le date degli allenamenti e la data della gara.
    Ora estrae sia paces che power_values e swim_paces indipendentemente dal tipo di sport.
    Utilizza estrazione diretta per le frequenze cardiache.
    
    Args:
        excel_file: Percorso del file Excel di input
        output_file: Percorso del file YAML di output (opzionale)
        sport_type: Tipo di sport (non più usato, viene ora estratto da ogni allenamento)
    """
    # Se non viene specificato un file di output, creiamo uno con lo stesso nome ma estensione .yaml
    if output_file is None:
        output_file = os.path.splitext(excel_file)[0] + '.yaml'
    
    print(f"Convertendo {excel_file} in {output_file}...")
    
    # Carica il file Excel
    try:
        # Leggi esplicitamente con le intestazioni nella seconda riga (header=1)
        df = pd.read_excel(excel_file, sheet_name='Workouts', header=1)
        
        # Verifica che ci siano le colonne richieste
        required_cols = ['Week', 'Session', 'Description', 'Steps']
        if all(col in df.columns for col in required_cols):
            print("Foglio 'Workouts' trovato con intestazioni nella seconda riga.")
        else:
            # Verifica se le colonne esistono ma con case diverso
            df_cols_lower = [col.lower() for col in df.columns]
            missing = []
            
            for req_col in required_cols:
                if req_col.lower() not in df_cols_lower:
                    missing.append(req_col)
            
            if missing:
                raise ValueError(f"Colonne mancanti nel foglio 'Workouts': {', '.join(missing)}")
            else:
                # Rinomina le colonne per uniformarle
                rename_map = {}
                for col in df.columns:
                    for req_col in required_cols:
                        if col.lower() == req_col.lower():
                            rename_map[col] = req_col
                
                df = df.rename(columns=rename_map)
                print("Colonne rinominate per uniformità.")
        
        # Ora puoi continuare con la lettura del resto del file
        xls = pd.ExcelFile(excel_file)
        
    except Exception as e:
        raise ValueError(f"Errore nel caricamento del foglio 'Workouts': {str(e)}")
    
    # Dizionario che conterrà il piano completo
    plan = {
        'config': {
            'margins': {
                'faster': '0:03',
                'slower': '0:03',
                'power_up': 10,
                'power_down': 10,
                'hr_up': 5,
                'hr_down': 5
            },
            'name_prefix': '',
        }
    }
    
    # Estrai il nome atleta dalla prima riga se presente
    try:
        athlete_row = pd.read_excel(excel_file, sheet_name='Workouts', header=None, nrows=1)
        athlete_text = str(athlete_row.iloc[0, 0])
        
        if athlete_text and athlete_text.strip().startswith("Atleta:"):
            athlete_name = athlete_text.replace("Atleta:", "").strip()
            if athlete_name:
                # Aggiungi il nome dell'atleta sia alla radice che nella sottosezione config
                plan['config']['athlete_name'] = athlete_name
                plan['athlete_name'] = athlete_name  # Aggiungi anche alla radice per compatibilità
                print(f"Nome atleta estratto: {athlete_name}")
    except Exception as e:
        print(f"Nota: impossibile estrarre il nome dell'atleta: {str(e)}")
    
    # Estrai la data della gara SOLO dal foglio Config
    race_day = None
    try:
        if 'Config' in xls.sheet_names:
            config_df = pd.read_excel(excel_file, sheet_name='Config')
            race_day_rows = config_df[config_df.iloc[:, 0] == 'race_day']
            
            if not race_day_rows.empty and pd.notna(race_day_rows.iloc[0, 1]):
                race_day_value = race_day_rows.iloc[0, 1]
                
                # Gestisci diversi formati di data
                if isinstance(race_day_value, datetime):
                    race_day = race_day_value.strftime("%Y-%m-%d")
                elif isinstance(race_day_value, str):
                    try:
                        # Prova a interpretare il formato
                        if len(race_day_value) == 10 and race_day_value[4] == '-' and race_day_value[7] == '-':
                            # Già in formato YYYY-MM-DD
                            race_day = race_day_value
                        else:
                            # Prova altre interpretazioni comuni
                            try:
                                date_obj = datetime.strptime(race_day_value, "%d/%m/%Y").date()
                                race_day = date_obj.strftime("%Y-%m-%d")
                            except ValueError:
                                try:
                                    date_obj = datetime.strptime(race_day_value, "%m/%d/%Y").date()
                                    race_day = date_obj.strftime("%Y-%m-%d")
                                except ValueError:
                                    print(f"Impossibile interpretare la data: {race_day_value}")
                    except:
                        print(f"Errore nel parsing della data: {race_day_value}")
                else:
                    # Gestisci altri tipi di dato, come date numeriche
                    try:
                        race_day = pd.to_datetime(race_day_value).strftime("%Y-%m-%d")
                    except:
                        print(f"Impossibile convertire il valore in data: {race_day_value}")
                
                if race_day:
                    plan['config']['race_day'] = race_day
                    print(f"Data della gara trovata nel foglio Config: {race_day}")
            else:
                print("Campo 'race_day' non trovato nel foglio Config o valore mancante")
        else:
            print("Foglio Config non trovato nel file Excel")
    except Exception as e:
        print(f"Errore nell'estrazione della data della gara: {str(e)}")
    
    # Estrai le informazioni di configurazione dal foglio Config
    if 'Config' in xls.sheet_names:
        config_df = pd.read_excel(excel_file, sheet_name='Config')
        
        # Estrai il prefisso del nome (se presente)
        name_prefix_rows = config_df[config_df.iloc[:, 0] == 'name_prefix']
        if not name_prefix_rows.empty and pd.notna(name_prefix_rows.iloc[0, 1]):
            # Assicurati che il prefisso termini con uno spazio
            prefix = str(name_prefix_rows.iloc[0, 1]).strip()
            # Aggiungi uno spazio alla fine se non c'è già
            if prefix and not prefix.endswith(' '):
                prefix = prefix + ' '
            plan['config']['name_prefix'] = prefix
        
        # Estrai i margini (se presenti)
        margins_rows = config_df[config_df.iloc[:, 0] == 'margins']
        if not margins_rows.empty:
            # Controlla se ci sono valori per i margini
            if pd.notna(margins_rows.iloc[0, 1]):
                plan['config']['margins']['faster'] = str(margins_rows.iloc[0, 1]).strip()
            if pd.notna(margins_rows.iloc[0, 2]):
                plan['config']['margins']['slower'] = str(margins_rows.iloc[0, 2]).strip()
            if pd.notna(margins_rows.iloc[0, 3]) and not pd.isna(margins_rows.iloc[0, 3]):
                try:
                    plan['config']['margins']['hr_up'] = int(margins_rows.iloc[0, 3])
                except (ValueError, TypeError):
                    print(f"Valore hr_up non valido: {margins_rows.iloc[0, 3]}")
            if pd.notna(margins_rows.iloc[0, 4]) and not pd.isna(margins_rows.iloc[0, 4]):
                try:
                    plan['config']['margins']['hr_down'] = int(margins_rows.iloc[0, 4])
                except (ValueError, TypeError):
                    print(f"Valore hr_down non valido: {margins_rows.iloc[0, 4]}")
            # Aggiungi anche power_up e power_down se presenti
            if pd.notna(margins_rows.iloc[0, 5]) and not pd.isna(margins_rows.iloc[0, 5]):
                try:
                    plan['config']['margins']['power_up'] = int(margins_rows.iloc[0, 5])
                except (ValueError, TypeError):
                    print(f"Valore power_up non valido: {margins_rows.iloc[0, 5]}")
            if pd.notna(margins_rows.iloc[0, 6]) and not pd.isna(margins_rows.iloc[0, 6]):
                try:
                    plan['config']['margins']['power_down'] = int(margins_rows.iloc[0, 6])
                except (ValueError, TypeError):
                    print(f"Valore power_down non valido: {margins_rows.iloc[0, 6]}")
        
        # Estrai preferred_days se presente
        preferred_days_rows = config_df[config_df.iloc[:, 0] == 'preferred_days']
        if not preferred_days_rows.empty and pd.notna(preferred_days_rows.iloc[0, 1]):
            preferred_days_value = str(preferred_days_rows.iloc[0, 1]).strip()
            plan['config']['preferred_days'] = preferred_days_value
        
        # Estrai athlete_name se presente
        athlete_name_rows = config_df[config_df.iloc[:, 0] == 'athlete_name']
        if not athlete_name_rows.empty and pd.notna(athlete_name_rows.iloc[0, 1]):
            athlete_name = str(athlete_name_rows.iloc[0, 1]).strip()
            if athlete_name:
                plan['config']['athlete_name'] = athlete_name
                plan['athlete_name'] = athlete_name  # Anche nella radice per compatibilità
                print(f"Nome atleta trovato nella configurazione: {athlete_name}")
    
    # Assicurati che il nome dell'atleta sia nella configurazione principale
    if 'athlete_name' in plan:
        plan['config']['athlete_name'] = plan['athlete_name'] 
    elif 'athlete_name' in plan['config']:
        plan['athlete_name'] = plan['config']['athlete_name']
    
    # MODIFICATO: Utilizziamo la nuova funzione di estrazione diretta per le frequenze cardiache
    heart_rates = extract_heart_rates_from_excel(excel_file)
    
    # Aggiungi le frequenze cardiache estratte, se presenti
    if heart_rates:
        plan['config']['heart_rates'] = heart_rates
        print(f"Frequenze cardiache estratte con successo: {heart_rates}")
    else:
        # Imposta un dizionario vuoto anche se non abbiamo trovato frequenze cardiache
        plan['config']['heart_rates'] = {}
        print("Nessuna frequenza cardiaca trovata o errore durante l'estrazione")
    
    # Estrai ritmi, velocità e passi vasca
    paces, swim_paces, power_values, _ = extract_paces_and_speeds_from_excel(excel_file)
    
    # Aggiungi direttamente al piano (non dentro config) SOLO se contengono valori
    if paces:
        plan['paces'] = paces
    if swim_paces:
        plan['swim_paces'] = swim_paces
    if power_values:
        plan['power_values'] = power_values
    
    # Dictionary to store workout descriptions for comments
    workout_descriptions = {}
    
    # Processa gli allenamenti dal DataFrame
    for _, row in df.iterrows():
        # Verifica che ci siano i dati necessari
        if pd.isna(row['Week']) or pd.isna(row['Session']) or pd.isna(row['Description']) or pd.isna(row['Steps']):
            continue
        
        # Estrai i dati
        week = str(int(row['Week'])).zfill(2)  # Formatta come 01, 02, ecc.
        session = str(int(row['Session'])).zfill(2)
        description = str(row['Description']).strip()
        
        # Salta le righe con description "athlete_name" (che potrebbero essere importate erroneamente)
        if description.lower() == "athlete_name" or "athlete_name" in description.lower():
            print(f"Ignorata riga con description '{description}' (non è un allenamento valido)")
            continue
        
        # Crea il nome completo dell'allenamento (senza includere la data)
        full_name = f"W{week}S{session} {description}"
        
        # Memorizza la descrizione per i commenti
        workout_descriptions[full_name] = description
        
        # Estrai i passi dell'allenamento
        steps_str = str(row['Steps']).strip()
        
        # Determina il tipo di sport
        workout_sport = "running"  # Default
        
        # Usa la colonna Sport se presente
        if 'Sport' in df.columns and pd.notna(row['Sport']):
            sport_value = str(row['Sport']).strip().lower()
            
            # Estrai il tipo di sport dal valore (rimuovendo eventuali emoji)
            if "running" in sport_value:
                workout_sport = "running"
            elif "cycling" in sport_value:
                workout_sport = "cycling"
            elif "swimming" in sport_value:
                workout_sport = "swimming"
        
        # Prepara la lista dei passi
        workout_steps = parse_workout_steps(steps_str, full_name, workout_sport)
        
        # Aggiungi metadati del tipo di sport come primo elemento 
        sport_type_meta = {"sport_type": workout_sport}
        workout_steps.insert(0, sport_type_meta)
        
        # Aggiungi la data come secondo elemento se disponibile
        if 'Date' in df.columns and pd.notna(row['Date']):
            date_value = row['Date']
            if isinstance(date_value, str):
                formatted_date = date_value
            else:
                # Se è un oggetto datetime o date, formattalo come stringa
                formatted_date = date_value.strftime("%Y-%m-%d") if hasattr(date_value, 'strftime') else str(date_value)
            
            # Aggiungi la data come secondo elemento dei passi
            date_step = {"date": formatted_date}
            workout_steps.insert(1, date_step)
        
        # Aggiungi l'allenamento al piano (senza la data nel nome)
        plan[full_name] = workout_steps
    
    # Salva il piano in formato YAML
    with open(output_file, 'w', encoding='utf-8') as f:
        # Usa NoAliasDumper per evitare riferimenti YAML
        yaml.dump(plan, f, default_flow_style=False, sort_keys=False, Dumper=NoAliasDumper)
    
    print(f"Conversione completata! File YAML salvato in: {output_file}")
    
    # Ora aggiungi i commenti al file
    add_comments_to_yaml(output_file, workout_descriptions)
    
    return plan


def main():
    """Main function for command line use"""
    # Define command line arguments
    parser = argparse.ArgumentParser(description='Convert an Excel file to a YAML file for garmin-planner')
    parser.add_argument('--excel', '-e', help='Path to the input Excel file', default='')
    parser.add_argument('--output', '-o', help='Path to the output YAML file (optional)')
    parser.add_argument('--create-sample', '-s', action='store_true', help='Create a sample Excel file')
    parser.add_argument('--sample-name', help='Name for the sample Excel file', default='sample_training_plan.xlsx')
    parser.add_argument('--sport-type', help='Type of sport (running or cycling)', choices=['running', 'cycling'], default='running')
    
    args = parser.parse_args()
    
    # Create a sample file if requested
    if args.create_sample:
        sample_file = create_sample_excel(args.sample_name, args.sport_type)
        if sample_file:
            # If specified --excel, immediately convert the sample file
            if args.excel == '':
                args.excel = sample_file
    
    # Verify that an input file is specified
    if not args.excel:
        logging.error("ERROR: You must specify an input Excel file (--excel)")
        logging.info("Use --create-sample to create a sample file")
        parser.print_help()
        return
    
    # Verify that the Excel file exists
    if not os.path.exists(args.excel):
        logging.error(f"ERROR: File {args.excel} does not exist")
        return
    
    # Convert the Excel file to YAML
    try:
        excel_to_yaml(args.excel, args.output, args.sport_type)
        logging.info("Operation completed successfully!")
    except Exception as e:
        logging.error(f"ERROR: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()